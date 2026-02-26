#!/usr/bin/env python3
"""
Vergelijk Pre-season teamindelingen (Excel) met spelerspaden.json.

Strategie: TERUGBLIK-EERST
  1. Terugblik (hoogste prioriteit): "Team vorig seizoen" uit volgend seizoen
  2. Vooruitblik (secundair): teamtoewijzing in huidig seizoen Pre-season
  3. Spelerspaden (fallback): behouden als geen Pre-season data

Gebruik:
    python scripts/vergelijk-preseason.py              # alleen rapport
    python scripts/vergelijk-preseason.py --patch      # rapport + correcties doorvoeren

Output:
    data/spelers/preseason-vergelijking.json
"""

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime

import openpyxl

# ── Hergebruik functies uit parse_spelerspaden ──────────────────────────
ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
sys.path.insert(0, os.path.join(ROOT, "data", "spelers"))

from parse_spelerspaden import (
    TUSSENVOEGSELS,
    _fuzzy_achternaam,
    extract_achternaam_voornaam,
    make_fuzzy_key,
    make_match_key,
    normalize_name,
    normalize_team,
)

# ── Constanten ──────────────────────────────────────────────────────────
PRESEASON_FILE = os.path.join(ROOT, "docs", "Pre-season teamindelingen overzicht.xlsx")
SPELERSPADEN_FILE = os.path.join(ROOT, "data", "spelers", "spelerspaden.json")
HERTELLING_FILE = os.path.join(ROOT, "data", "spelers", "hertelling.json")
OUTPUT_FILE = os.path.join(ROOT, "data", "spelers", "preseason-vergelijking.json")

OVERLAP_SEIZOENEN = [
    "2020-2021", "2021-2022", "2022-2023",
    "2023-2024", "2024-2025", "2025-2026",
]

# Combinatie-teams: als Pre-season een pool-team noemt en spelerspaden een
# specifiek team uit die pool, is dat geen echte mismatch.
COMBINED_TEAMS = {
    "S1S2": {"S1", "S2", "S1S2"},
    "A1A2": {"A1", "A2", "A1A2"},
    "B1B2": {"B1", "B2", "B1B2"},
    "S5S6S7": {"S5", "S6", "S7", "S5S6", "S5S6S7", "S5-6-7", "S567"},
    "S3S4": {"S3", "S4", "S3S4"},
    "U17-1U17-2": {"U17", "U17-1", "U17-2", "U17-1U17-2"},
    "U19-1U19-2": {"U19", "U19-1", "U19-2", "U19-1U19-2"},
}

# Termen in "Team vorig seizoen" die geen echt team zijn
TERUGBLIK_SKIP = {"stopt", "nieuw", "gestopt", "-", "stop", "nieuwe", "n", "onbekend", ""}

# Woorden/patronen die aangeven dat een rij geen speler is
NON_PLAYER_PATTERNS = [
    r"^U-?\d+\s*/\s*\w+$",
    r"^(?:Oranje|Rood|Geel|Groen|Blauw)\s*$",
    r"(?:verder worden besproken|onderlinge verhouding|aan te vullen)",
    r"^Voorheen\b",
    r"(?:teams?\s+)?spelen\b",
    r"^heer\s+uit\b",
    r"^Nog\s+aan\s+te\b",
    r"korfbal\s+met\s+een",
    r"leeftijdsverschil\b",
    r"gewisseld\s+op\s+tijd",
    r"sociaal\s+ingedeeld",
    r"jongste\s+en\s+oudste",
    r"hoge\s+korven",
    r"niet\s+meer\s+dan\b",
    r"^\w+\s+kwart\b",
    r"^Er\s+wordt\b",
    r"^De\s+teams\b",
    r"^Het\s+leeftijds",
    r"^de\s+jongste\b",
]

# Handmatige correcties: ouder/kind fouten zonder terugblik (2021-2022)
HANDMATIGE_CORRECTIES = {
    # speler_id → team
    # Nina Hulsker: 3 jaar S6, dan D3 (jeugd) — onmogelijk
    "Nina Hulsker|V|2021-2022": "S6",
    # Suzanne Klink: jarenlang A3, dan E6 (jong kind) — ouder/kind
    "Suzanne Klink|V|2021-2022": "S7",
    # Mandy Dullemeijer: was F2, dan C2 — onmogelijk
    "Mandy Dullemeijer|V|2021-2022": "F2",
    # Erinn Bax: was S5, dan A1 — onwaarschijnlijk
    "Erinn Bax|V|2021-2022": "S5",
    # Dilan Canli: was A4, PS=S5S6S7
    "Dilan Canli|V|2021-2022": "S5",
}


# ── Pre-season teamnaam normalisatie ────────────────────────────────────
def normalize_preseason_team(raw):
    """Normaliseer een Pre-season teamnaam naar een code vergelijkbaar met spelerspaden."""
    if not raw:
        return None
    t = raw.strip()

    # Strip "Selectie " prefix
    t = re.sub(r'^Selectie\s+', '', t, flags=re.IGNORECASE)
    # Strip "Jong " prefix
    t = re.sub(r'^Jong\s+', '', t, flags=re.IGNORECASE)
    # Strip "Oranje Wit " / "OW " prefix
    t = re.sub(r'^(?:Oranje\s+Wit|OW)\s*', '', t, flags=re.IGNORECASE).strip()

    # Als er een parenthesized team-code is: "3 (S3)" → "S3", "(S1/S2)" → "S1S2"
    m = re.match(r'^.*?\(([^)]+)\)\s*$', t)
    if m:
        inner = m.group(1).strip()
        inner = re.sub(r'\s+en\s+', '/', inner, flags=re.IGNORECASE)
        codes = [c.strip() for c in inner.split('/') if c.strip()]
        codes = [c for c in codes if re.match(r'^[A-Z]\w*\d[\w-]*$', c, re.IGNORECASE)]
        if len(codes) == 1:
            return codes[0]
        elif len(codes) > 1:
            return "".join(codes)

    # "midweek" → "MW1"
    if t.lower().startswith("midweek"):
        return "MW1"

    # Als er alleen een getal overblijft → seniorenteam
    if re.match(r'^\d+$', t):
        return f"S{t}"

    # A-selectie / B-selectie zonder parens
    m2 = re.match(r'^([A-Z])-selectie$', t, re.IGNORECASE)
    if m2:
        letter = m2.group(1).upper()
        return f"{letter}1{letter}2"

    # U-selectie patronen: "U17-Selectie" → "U17"
    m3 = re.match(r'^(U\d+)-?[Ss]electie$', t)
    if m3:
        return m3.group(1)

    # Normaliseer via bestaande functie als fallback
    result = normalize_team(t)
    return result if result else t


def normalize_geslacht(raw):
    """Normaliseer geslacht naar M/V."""
    g = raw.strip().upper() if raw else ""
    if g in ("DAMES", "VROUW", "D"):
        return "V"
    elif g in ("HEREN", "MAN", "H"):
        return "M"
    return g


def is_non_player_row(speler_naam):
    """Filter rijen die geen echte spelersnamen zijn."""
    if not speler_naam or len(speler_naam.strip()) < 3:
        return True
    s = speler_naam.strip()
    for pat in NON_PLAYER_PATTERNS:
        if re.search(pat, s, re.IGNORECASE):
            return True
    words = s.split()
    if len(words) > 4:
        return True
    if s.endswith(".") and not re.match(r'^.+\s+[A-Z]\.$', s):
        return True
    return False


def is_pool_assignment(team_raw):
    """Check of een Pre-season team een pool-toewijzing is (bijv. "E1/E2/E3/E4/E5")."""
    return team_raw and "/" in team_raw and len(team_raw.split("/")) > 2


def teams_equivalent(team_a, team_b):
    """Check of twee teams equivalent zijn (inclusief combinatie-teams)."""
    if not team_a or not team_b:
        return False
    if team_a == team_b:
        return True
    for combined, members in COMBINED_TEAMS.items():
        if team_a in members and team_b in members:
            return True
        if team_a == combined and team_b in members:
            return True
        if team_b == combined and team_a in members:
            return True
    return False


# ── Match-index bouwen ──────────────────────────────────────────────────
def build_match_index(spelers):
    """Bouw een index per seizoen × geslacht met meerdere key-varianten."""
    index = defaultdict(lambda: defaultdict(dict))

    for sp in spelers:
        geslacht = sp.get("geslacht", "")
        if not geslacht:
            continue

        ach, vn = extract_achternaam_voornaam(sp["naam"])
        roepnaam = sp.get("roepnaam", "")

        keys = set()

        # Exact keys
        for v in [vn, roepnaam]:
            if v:
                k = make_match_key(ach, v)
                if k and k != "|":
                    keys.add(k)

        # Root keys
        ach_parts = ach.split() if ach else []
        root_parts = [p for p in ach_parts if p.lower() not in TUSSENVOEGSELS]
        root = " ".join(root_parts)
        if root:
            for v in [vn, roepnaam]:
                if v:
                    first_word = v.split()[0] if v.split() else ""
                    if first_word and first_word.lower() not in TUSSENVOEGSELS:
                        k = make_match_key(root, first_word)
                        if k and k != "|":
                            keys.add(k)

        # Hyphenated achternaam
        if ach and "-" in ach:
            first_part = ach.split("-")[0].strip()
            if first_part:
                for v in [vn, roepnaam]:
                    if v:
                        first_word = v.split()[0] if v.split() else v
                        if first_word and first_word.lower() not in TUSSENVOEGSELS:
                            k = make_match_key(first_part, first_word)
                            if k and k != "|":
                                keys.add(k)

        # Fuzzy keys
        for v in [vn, roepnaam]:
            if v:
                fk = make_fuzzy_key(ach, v)
                if fk:
                    keys.add(fk)
                if root and root != ach:
                    fk2 = make_fuzzy_key(root, v)
                    if fk2:
                        keys.add(fk2)

        for seizoen in sp.get("seizoenen", {}):
            if seizoen in OVERLAP_SEIZOENEN:
                for key in keys:
                    if key not in index[seizoen][geslacht]:
                        index[seizoen][geslacht][key] = sp

    return index


def match_player(voornaam, achternaam, seizoen, geslacht, index):
    """Probeer een Pre-season speler te matchen in de index."""
    idx = index.get(seizoen, {}).get(geslacht, {})
    if not idx:
        return None, None

    candidates = []

    k1 = make_match_key(achternaam, voornaam)
    if k1 and k1 != "|":
        candidates.append((k1, "exact"))

    ach_parts = achternaam.split() if achternaam else []
    root_parts = [p for p in ach_parts if p.lower() not in TUSSENVOEGSELS]
    root = " ".join(root_parts)
    if root:
        k2 = make_match_key(root, voornaam)
        if k2 and k2 != "|" and k2 != k1:
            candidates.append((k2, "root"))
        first_word = voornaam.split()[0] if voornaam and voornaam.split() else ""
        if first_word and first_word.lower() not in TUSSENVOEGSELS:
            k3 = make_match_key(root, first_word)
            if k3 and k3 != "|" and k3 not in [c[0] for c in candidates]:
                candidates.append((k3, "root_first"))

    if voornaam:
        first_word = voornaam.split()[0] if voornaam.split() else ""
        if first_word and first_word.lower() not in TUSSENVOEGSELS:
            k4 = make_match_key(achternaam, first_word)
            if k4 and k4 != "|" and k4 not in [c[0] for c in candidates]:
                candidates.append((k4, "first_word"))

    fk = make_fuzzy_key(achternaam, voornaam)
    if fk:
        candidates.append((fk, "fuzzy"))
    if root and root != achternaam:
        fk2 = make_fuzzy_key(root, voornaam)
        if fk2 and fk2 != fk:
            candidates.append((fk2, "fuzzy_root"))

    for key, match_type in candidates:
        if key in idx:
            return idx[key], match_type

    return None, None


# ── Terugblik-index ─────────────────────────────────────────────────────
def build_terugblik_index(preseason_rows, match_index):
    """Bouw terugblik-index: (speler_id, vorig_seizoen) → team.

    Gebruikt de "Team vorig seizoen" kolom uit het VOLGENDE seizoen.
    """
    terugblik = {}
    stats = {"gevonden": 0, "overgeslagen_skip": 0, "niet_gematcht": 0}

    for row in preseason_rows:
        tv = row.get("team_vorig", "").strip()
        if not tv or tv.lower() in TERUGBLIK_SKIP:
            stats["overgeslagen_skip"] += 1
            continue
        if row["seizoen"] == "2020-2021":
            continue

        jaar1 = int(row["seizoen"].split("-")[0])
        vorig_seizoen = f"{jaar1 - 1}-{jaar1}"
        if vorig_seizoen not in OVERLAP_SEIZOENEN:
            continue

        naam = row["speler"]
        if is_non_player_row(naam):
            continue
        ach, vn = extract_achternaam_voornaam(naam)
        geslacht = normalize_geslacht(row["geslacht"])

        speler, _ = match_player(vn, ach, vorig_seizoen, geslacht, match_index)
        if not speler:
            stats["niet_gematcht"] += 1
            continue

        team = normalize_preseason_team(tv)
        if not team:
            team = tv

        key = (speler["speler_id"], vorig_seizoen)
        # Eerste terugblik wint (geen overschrijving)
        if key not in terugblik:
            terugblik[key] = team
            stats["gevonden"] += 1

    return terugblik, stats


# ── Excel lezen ─────────────────────────────────────────────────────────
def read_preseason_excel(path):
    """Lees de 'Overzicht' sheet en return een lijst van dicts."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Overzicht"]

    raw_rows = []
    header = None
    for row in ws.iter_rows(values_only=True):
        cells = [str(c).strip() if c else "" for c in row]
        if header is None:
            header = cells
            continue
        if not cells[0]:
            continue
        raw_rows.append({
            "seizoen": cells[0],
            "team_raw": cells[1],
            "categorie": cells[2],
            "speler": cells[3],
            "geslacht": cells[4] if len(cells) > 4 else "",
            "team_vorig": cells[5] if len(cells) > 5 else "",
        })
    wb.close()

    # Merge split-namen
    rows = []
    i = 0
    while i < len(raw_rows):
        r = raw_rows[i]
        naam = r["speler"]
        while (naam.endswith("-") and i + 1 < len(raw_rows)
               and raw_rows[i + 1]["seizoen"] == r["seizoen"]
               and raw_rows[i + 1]["team_raw"] == r["team_raw"]):
            i += 1
            naam = naam + raw_rows[i]["speler"]
            if raw_rows[i]["geslacht"] and not r["geslacht"]:
                r["geslacht"] = raw_rows[i]["geslacht"]
            if raw_rows[i]["team_vorig"] and not r["team_vorig"]:
                r["team_vorig"] = raw_rows[i]["team_vorig"]
        r["speler"] = naam

        if not is_non_player_row(naam):
            rows.append(r)
        i += 1

    # Lees Samenvatting voor sanity check
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2["Samenvatting"]
    samenvatting = {}
    header2 = None
    for row in ws2.iter_rows(values_only=True):
        cells = [str(c).strip() if c else "" for c in row]
        if header2 is None:
            header2 = cells
            continue
        if cells[0]:
            samenvatting[cells[0]] = {
                "teams": int(cells[1]) if cells[1] else 0,
                "totaal": int(cells[2]) if cells[2] else 0,
                "dames": int(cells[3]) if cells[3] else 0,
                "heren": int(cells[4]) if cells[4] else 0,
            }
    wb2.close()

    return rows, samenvatting


# ── Hoofdlogica ─────────────────────────────────────────────────────────
def run_comparison():
    """Voer de volledige vergelijking uit met terugblik-eerst strategie."""

    # Laad bronnen
    print("Bronnen laden...")
    preseason_rows, samenvatting = read_preseason_excel(PRESEASON_FILE)
    print(f"  Pre-season: {len(preseason_rows)} rijen, {len(samenvatting)} seizoenen")

    with open(SPELERSPADEN_FILE, "r", encoding="utf-8") as f:
        spelers = json.load(f)
    print(f"  Spelerspaden: {len(spelers)} spelers")

    # Bouw match-index
    print("Match-index bouwen...")
    index = build_match_index(spelers)

    speler_by_id = {sp["speler_id"]: sp for sp in spelers}

    # ── Terugblik-index bouwen ────────────────────────────────────────
    print("Terugblik-index bouwen...")
    terugblik, tb_stats = build_terugblik_index(preseason_rows, index)
    print(f"  Terugblik: {tb_stats['gevonden']} entries, "
          f"{tb_stats['niet_gematcht']} niet gematcht")

    # ── Vergelijking ──────────────────────────────────────────────────
    print("Vergelijking uitvoeren...\n")

    results = {seizoen: {
        "rijen": 0,
        "matched": 0,
        "exact_match": 0,
        "normalization_match": 0,
        "generic_vs_specific": 0,
        "mismatch": 0,
        "missing_team_in_spelerspaden": 0,
        "missing_in_spelerspaden": 0,
        "vorig_match": 0,
        "vorig_mismatch": 0,
        "vorig_no_prev": 0,
        "vorig_checked": 0,
        "mismatches": [],
        "missing": [],
        "vorig_mismatches": [],
        "generic": [],
        "all_matched": [],  # Alle gematchte spelers met hun Pre-season team
    } for seizoen in OVERLAP_SEIZOENEN}

    team_norm_log = Counter()
    matched_speler_seasons = set()

    for row in preseason_rows:
        seizoen = row["seizoen"]
        if seizoen not in results:
            continue

        r = results[seizoen]
        r["rijen"] += 1

        naam_raw = row["speler"]
        ach, vn = extract_achternaam_voornaam(naam_raw)

        team_norm = normalize_preseason_team(row["team_raw"])
        team_norm_log[f"{row['team_raw']} → {team_norm}"] += 1

        geslacht = normalize_geslacht(row["geslacht"])

        speler, match_type = match_player(vn, ach, seizoen, geslacht, index)

        if not speler:
            r["missing_in_spelerspaden"] += 1
            r["missing"].append({
                "speler": naam_raw,
                "geslacht": geslacht,
                "preseason_team": team_norm,
                "preseason_team_raw": row["team_raw"],
            })
            continue

        r["matched"] += 1
        matched_speler_seasons.add((speler["speler_id"], seizoen))

        # Bewaar voor correctie-berekening
        r["all_matched"].append({
            "speler": naam_raw,
            "speler_id": speler["speler_id"],
            "geslacht": geslacht,
            "match_type": match_type,
            "preseason_team": team_norm,
            "preseason_team_raw": row["team_raw"],
        })

        # Vergelijk team (voor rapportage)
        sp_team = speler.get("seizoenen", {}).get(seizoen, {}).get("team", "")

        if not sp_team:
            r["missing_team_in_spelerspaden"] += 1
        elif team_norm == sp_team:
            r["exact_match"] += 1
        else:
            equiv = teams_equivalent(team_norm, sp_team)
            if equiv:
                r["normalization_match"] += 1
            elif sp_team in ("Senioren",):
                r["generic_vs_specific"] += 1
                r["generic"].append({
                    "speler": naam_raw,
                    "speler_id": speler["speler_id"],
                    "preseason_team": team_norm,
                    "spelerspaden_team": sp_team,
                })
            else:
                r["mismatch"] += 1
                r["mismatches"].append({
                    "speler": naam_raw,
                    "speler_id": speler["speler_id"],
                    "geslacht": geslacht,
                    "match_type": match_type,
                    "preseason_team": team_norm,
                    "preseason_team_raw": row["team_raw"],
                    "spelerspaden_team": sp_team,
                })

        # "Team vorig seizoen" validatie (voor rapportage)
        if row["team_vorig"]:
            tv = row["team_vorig"].strip()
            if tv and tv.lower() not in TERUGBLIK_SKIP:
                jaar1 = int(seizoen.split("-")[0])
                vorig_seizoen = f"{jaar1 - 1}-{jaar1}"
                vorig_data = speler.get("seizoenen", {}).get(vorig_seizoen)
                if vorig_data:
                    sp_vorig = vorig_data.get("team", "")
                    vorig_norm = normalize_preseason_team(tv) or tv
                    r["vorig_checked"] += 1
                    if teams_equivalent(vorig_norm, sp_vorig):
                        r["vorig_match"] += 1
                    else:
                        r["vorig_mismatch"] += 1
                        r["vorig_mismatches"].append({
                            "speler": naam_raw,
                            "speler_id": speler["speler_id"],
                            "preseason_vorig": tv,
                            "preseason_vorig_norm": vorig_norm,
                            "spelerspaden_vorig": sp_vorig,
                            "vorig_seizoen": vorig_seizoen,
                        })
                elif vorig_seizoen in OVERLAP_SEIZOENEN:
                    r["vorig_no_prev"] += 1

    # ── Reverse check ─────────────────────────────────────────────────
    missing_in_preseason = {seizoen: [] for seizoen in OVERLAP_SEIZOENEN}
    for sp in spelers:
        for seizoen in OVERLAP_SEIZOENEN:
            if seizoen in sp.get("seizoenen", {}):
                if (sp["speler_id"], seizoen) not in matched_speler_seasons:
                    team = sp["seizoenen"][seizoen].get("team", "")
                    missing_in_preseason[seizoen].append({
                        "speler_id": sp["speler_id"],
                        "naam": sp["naam"],
                        "roepnaam": sp.get("roepnaam", ""),
                        "geslacht": sp.get("geslacht", ""),
                        "spelerspaden_team": team,
                    })

    # ── Correcties: TERUGBLIK-EERST ───────────────────────────────────
    correcties = []
    corrected_keys = set()  # (speler_id, seizoen) al gecorrigeerd
    corr_stats = {"terugblik": 0, "vooruitblik": 0, "handmatig": 0, "skipped_pool": 0}

    # Pass 1: Terugblik — loop de VOLLEDIGE terugblik-index door
    # Dit vangt ook spelers op die niet in de Pre-season van dat seizoen staan
    for (speler_id, seizoen), tb_team in terugblik.items():
        sp = speler_by_id.get(speler_id)
        if not sp or seizoen not in sp["seizoenen"]:
            continue
        current_team = sp["seizoenen"][seizoen].get("team", "")
        if not teams_equivalent(tb_team, current_team):
            correcties.append({
                "speler_id": speler_id,
                "seizoen": seizoen,
                "veld": "team",
                "oud": current_team,
                "nieuw": tb_team,
                "bron": "terugblik",
                "speler_naam": sp["naam"],
            })
            corr_stats["terugblik"] += 1
        corrected_keys.add((speler_id, seizoen))

    # Pass 2: Vooruitblik — alleen voor spelers ZONDER terugblik
    for seizoen in OVERLAP_SEIZOENEN:
        r = results[seizoen]
        for m in r["all_matched"]:
            key = (m["speler_id"], seizoen)
            if key in corrected_keys:
                continue  # Terugblik heeft voorrang

            sp = speler_by_id.get(m["speler_id"])
            if not sp or seizoen not in sp["seizoenen"]:
                continue

            ps_raw = m.get("preseason_team_raw", "")
            if is_pool_assignment(ps_raw):
                corr_stats["skipped_pool"] += 1
                continue

            current_team = sp["seizoenen"][seizoen].get("team", "")
            ps_team = m["preseason_team"]
            if ps_team and not teams_equivalent(ps_team, current_team):
                correcties.append({
                    "speler_id": m["speler_id"],
                    "seizoen": seizoen,
                    "veld": "team",
                    "oud": current_team,
                    "nieuw": ps_team,
                    "bron": "vooruitblik",
                    "speler_naam": m["speler"],
                })
                corr_stats["vooruitblik"] += 1
            corrected_keys.add(key)

    # Handmatige correcties (ouder/kind zonder terugblik)
    for key, nieuw_team in HANDMATIGE_CORRECTIES.items():
        naam, geslacht, seizoen = key.split("|")
        ach, vn = extract_achternaam_voornaam(naam)
        sp, _ = match_player(vn, ach, seizoen, geslacht, index)
        if sp and seizoen in sp["seizoenen"]:
            current = sp["seizoenen"][seizoen].get("team", "")
            if current != nieuw_team:
                if (sp["speler_id"], seizoen) not in corrected_keys:
                    correcties.append({
                        "speler_id": sp["speler_id"],
                        "seizoen": seizoen,
                        "veld": "team",
                        "oud": current,
                        "nieuw": nieuw_team,
                        "bron": "handmatig",
                        "speler_naam": naam,
                    })
                    corr_stats["handmatig"] += 1

    # Lowercase fixes (2014-2015)
    for sp in spelers:
        for seizoen, data in sp["seizoenen"].items():
            team = data.get("team", "")
            if team and team != team.upper() and re.match(r'^[a-z]\d+$', team):
                correcties.append({
                    "speler_id": sp["speler_id"],
                    "seizoen": seizoen,
                    "veld": "team",
                    "oud": team,
                    "nieuw": team.upper(),
                    "bron": "lowercase_fix",
                    "speler_naam": sp["naam"],
                })

    # ── Samenvatting ──────────────────────────────────────────────────
    totaal = {
        "rijen": 0, "matched": 0, "exact_match": 0,
        "normalization_match": 0, "generic_vs_specific": 0,
        "mismatch": 0, "missing_team_in_spelerspaden": 0,
        "missing_in_spelerspaden": 0,
        "missing_in_preseason": 0,
        "vorig_checked": 0, "vorig_match": 0, "vorig_mismatch": 0,
        "vorig_no_prev": 0,
        "correcties_totaal": len(correcties),
        "correcties_terugblik": corr_stats["terugblik"],
        "correcties_vooruitblik": corr_stats["vooruitblik"],
        "correcties_handmatig": corr_stats["handmatig"],
        "skipped_pools": corr_stats["skipped_pool"],
    }
    for seizoen in OVERLAP_SEIZOENEN:
        r = results[seizoen]
        for k in ["rijen", "matched", "exact_match", "normalization_match",
                   "generic_vs_specific", "mismatch", "missing_team_in_spelerspaden",
                   "missing_in_spelerspaden", "vorig_checked", "vorig_match",
                   "vorig_mismatch", "vorig_no_prev"]:
            totaal[k] += r[k]
        totaal["missing_in_preseason"] += len(missing_in_preseason[seizoen])

    # ── Console output ────────────────────────────────────────────────
    print("=" * 80)
    print("  Vergelijking Pre-season teamindelingen vs Spelerspaden")
    print("  Strategie: TERUGBLIK-EERST")
    print("=" * 80)
    print()

    header = f"{'Seizoen':<12} {'Rijen':>6} {'Match':>6} {'Exact':>6} {'Norm':>5} {'Gen.':>5} {'Mis':>5} {'Afwezig':>8} {'←Niet PS':>9}"
    print(header)
    print("-" * len(header))

    for seizoen in OVERLAP_SEIZOENEN:
        r = results[seizoen]
        mip = len(missing_in_preseason[seizoen])
        print(f"  {seizoen:<10} {r['rijen']:>6} {r['matched']:>6} "
              f"{r['exact_match']:>6} {r['normalization_match']:>5} "
              f"{r['generic_vs_specific']:>5} {r['mismatch']:>5} "
              f"{r['missing_in_spelerspaden']:>8} {mip:>9}")

    print("-" * len(header))
    print(f"  {'TOTAAL':<10} {totaal['rijen']:>6} {totaal['matched']:>6} "
          f"{totaal['exact_match']:>6} {totaal['normalization_match']:>5} "
          f"{totaal['generic_vs_specific']:>5} {totaal['mismatch']:>5} "
          f"{totaal['missing_in_spelerspaden']:>8} {totaal['missing_in_preseason']:>9}")

    if totaal["rijen"] > 0:
        pct = totaal["matched"] / totaal["rijen"] * 100
        pct_ex = totaal["exact_match"] / totaal["rijen"] * 100
        print(f"\n  Match-rate: {pct:.1f}% | Exact: {pct_ex:.1f}%")

    # Vorig seizoen
    print(f"\n  Vorig seizoen validatie: {totaal['vorig_checked']} gecontroleerd, "
          f"{totaal['vorig_match']} match, {totaal['vorig_mismatch']} mismatch")

    # Correcties
    print(f"\n  Correcties ({len(correcties)} totaal):")
    print(f"    Terugblik:    {corr_stats['terugblik']:>4}")
    print(f"    Vooruitblik:  {corr_stats['vooruitblik']:>4}")
    print(f"    Handmatig:    {corr_stats['handmatig']:>4}")
    print(f"    Pools bewaard:{corr_stats['skipped_pool']:>4}")

    # Sanity check
    print(f"\n  Sanity check (Samenvatting-sheet):")
    for seizoen in OVERLAP_SEIZOENEN:
        r = results[seizoen]
        if seizoen in samenvatting:
            sv = samenvatting[seizoen]
            delta = r["rijen"] - sv["totaal"]
            mark = "✓" if delta == 0 else f"Δ{delta:+d}"
            print(f"    {seizoen}: {r['rijen']} vs {sv['totaal']} {mark}")

    # Top correcties per bron
    for bron in ["terugblik", "vooruitblik", "handmatig"]:
        bron_corr = [c for c in correcties if c["bron"] == bron]
        if bron_corr:
            print(f"\n  Correcties [{bron}] (top 15 van {len(bron_corr)}):")
            for c in bron_corr[:15]:
                print(f"    {c['speler_naam']:<28} {c['seizoen']}  "
                      f"{c['oud']:<8} → {c['nieuw']}")
            if len(bron_corr) > 15:
                print(f"    ... en {len(bron_corr) - 15} meer")

    # Mismatches (na correctie zouden deze kleiner zijn)
    for seizoen in OVERLAP_SEIZOENEN:
        r = results[seizoen]
        if r["mismatches"]:
            print(f"\n  Directe mismatches {seizoen} ({len(r['mismatches'])}):")
            for m in r["mismatches"][:5]:
                print(f"    {m['speler']:<28} PS:{m['preseason_team']:<8} SP:{m['spelerspaden_team']:<8}")
            if len(r["mismatches"]) > 5:
                print(f"    ... en {len(r['mismatches']) - 5} meer")

    # Vorig seizoen mismatches
    all_vorig = []
    for seizoen in OVERLAP_SEIZOENEN:
        all_vorig.extend(results[seizoen]["vorig_mismatches"])
    if all_vorig:
        print(f"\n  Vorig seizoen mismatches (top 10 van {len(all_vorig)}):")
        for v in all_vorig[:10]:
            print(f"    {v['speler']:<28} {v['vorig_seizoen']}  "
                  f"PS:{v['preseason_vorig']:<8} SP:{v['spelerspaden_vorig']:<8}")

    # ── JSON rapport ──────────────────────────────────────────────────
    rapport = {
        "_meta": {
            "gegenereerd": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "strategie": "terugblik-eerst",
            "preseason_bestand": os.path.relpath(PRESEASON_FILE, ROOT),
            "spelerspaden_bestand": os.path.relpath(SPELERSPADEN_FILE, ROOT),
            "seizoenen": OVERLAP_SEIZOENEN,
        },
        "samenvatting": totaal,
        "samenvatting_sheet": samenvatting,
        "terugblik_stats": tb_stats,
        "correctie_stats": corr_stats,
        "per_seizoen": {},
        "correcties": correcties,
        "team_normalisatie": [
            {"raw_to_norm": k, "count": v}
            for k, v in sorted(team_norm_log.items(), key=lambda x: -x[1])
        ],
    }

    for seizoen in OVERLAP_SEIZOENEN:
        r = results[seizoen]
        rapport["per_seizoen"][seizoen] = {
            "rijen": r["rijen"],
            "matched": r["matched"],
            "exact_match": r["exact_match"],
            "normalization_match": r["normalization_match"],
            "generic_vs_specific": r["generic_vs_specific"],
            "mismatch": r["mismatch"],
            "missing_in_spelerspaden": r["missing_in_spelerspaden"],
            "missing_in_preseason": len(missing_in_preseason[seizoen]),
            "vorig_checked": r["vorig_checked"],
            "vorig_match": r["vorig_match"],
            "vorig_mismatch": r["vorig_mismatch"],
            "mismatches": r["mismatches"],
            "missing": r["missing"],
            "missing_in_preseason_lijst": missing_in_preseason[seizoen],
            "vorig_mismatches": r["vorig_mismatches"],
            "generic": r["generic"],
        }

    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(rapport, f, ensure_ascii=False, indent=2)
    print(f"\n  Rapport: {os.path.relpath(OUTPUT_FILE, ROOT)}")

    return spelers, correcties, rapport


# ── Patch-modus ─────────────────────────────────────────────────────────
def patch_spelerspaden(spelers, correcties):
    """Pas correcties toe op spelerspaden en herbereken hertelling."""
    speler_by_id = {sp["speler_id"]: sp for sp in spelers}
    patched = 0
    by_bron = Counter()

    for corr in correcties:
        sp = speler_by_id.get(corr["speler_id"])
        if not sp:
            continue
        seizoen = corr["seizoen"]
        if seizoen in sp["seizoenen"]:
            sp["seizoenen"][seizoen]["team"] = corr["nieuw"]
            patched += 1
            by_bron[corr["bron"]] += 1

    print(f"\n  Patch: {patched} correcties toegepast")
    for bron, count in sorted(by_bron.items()):
        print(f"    {bron}: {count}")

    with open(SPELERSPADEN_FILE, "w", encoding="utf-8") as f:
        json.dump(spelers, f, ensure_ascii=False, indent=2)
    print(f"  Geschreven: {os.path.relpath(SPELERSPADEN_FILE, ROOT)}")

    # Herbereken hertelling
    hertelling = {}
    for sp in spelers:
        for seizoen, data in sp["seizoenen"].items():
            if seizoen not in hertelling:
                hertelling[seizoen] = {"totaal": 0, "M": 0, "V": 0, "teams": {}, "aantal_teams": 0}
            h = hertelling[seizoen]
            h["totaal"] += 1
            g = sp.get("geslacht", "")
            if g in ("M", "V"):
                h[g] += 1
            team = data.get("team", "")
            if team:
                if team not in h["teams"]:
                    h["teams"][team] = {"M": 0, "V": 0}
                if g in ("M", "V"):
                    h["teams"][team][g] += 1

    for seizoen in hertelling:
        hertelling[seizoen]["aantal_teams"] = len(hertelling[seizoen]["teams"])

    hertelling_sorted = dict(sorted(hertelling.items()))

    with open(HERTELLING_FILE, "w", encoding="utf-8") as f:
        json.dump(hertelling_sorted, f, ensure_ascii=False, indent=2)
    print(f"  Geschreven: {os.path.relpath(HERTELLING_FILE, ROOT)}")


# ── Main ────────────────────────────────────────────────────────────────
def main():
    # Fix Windows console encoding
    if sys.platform == "win32":
        sys.stdout.reconfigure(encoding="utf-8")

    parser = argparse.ArgumentParser(
        description="Vergelijk Pre-season teamindelingen met spelerspaden.json (terugblik-eerst)"
    )
    parser.add_argument("--patch", action="store_true",
                        help="Pas correcties toe op spelerspaden.json + hertelling.json")
    args = parser.parse_args()

    spelers, correcties, rapport = run_comparison()

    if args.patch and correcties:
        patch_spelerspaden(spelers, correcties)
    elif args.patch:
        print("\n  Geen correcties om toe te passen.")

    print("\n  Klaar.")


if __name__ == "__main__":
    main()
