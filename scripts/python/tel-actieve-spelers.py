"""
tel-actieve-spelers.py

Telt het aantal unieke ingedeelde korfbalspelers per seizoen,
rechtstreeks uit de bronbestanden (Excel teamindelingen).

Scope: competitie + midweekteams. Exclusief: recreanten, AR-lijst, staf, placeholders.
Output: data/aggregaties/actieve-spelers-per-seizoen.json + stdout samenvatting.

Gebruik: python scripts/tel-actieve-spelers.py
"""

import json
import os
import re
from datetime import date

# --- Config ---

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
TEAMINDELINGEN_DIR = os.path.join(ROOT, "docs", "teamindelingen")
OUTPUT_PATH = os.path.join(ROOT, "data", "aggregaties", "actieve-spelers-per-seizoen.json")
TELLING_OUTPUT_PATH = os.path.join(ROOT, "data", "aggregaties", "teamindelingen-telling.json")
HERTELLING_PATH = os.path.join(ROOT, "data", "spelers", "hertelling.json")
TELLING_BESTAND = "Telling spelers per seizoen 2010-2019.xlsx"

BESTANDEN = [
    ("2010-2011", "Opstelling 20102011.xls", "opstellingslijst"),
    ("2011-2012", "Opstelling 20112012.xls", "opstellingslijst"),
    ("2012-2013", "Opstelling 20122013.xlsx", "opstellingslijst"),
    ("2013-2014", "opstelling 20132014.xls", "opstellingslijst"),
    ("2014-2015", "opstelling 20142015.xls", "opstellingslijst"),
    ("2015-2016", "opstelling 20152016.xls", "opstellingslijst"),
    ("2016-2017", "Opstelling 20162017.xlsx", "opstellingslijst"),
    ("2017-2018", "Opstelling 20172018.xlsx", "opstellingslijst"),
    ("2018-2019", "A2 zwt - Oranje Wit (D) 2018 2019.xlsm", "a2_formulier"),
    ("2019-2020", "Opstelling 20192020.xlsx", "opstellingslijst"),
    ("2020-2021", "Teamindeling seizoen 20-21.xlsx", "teamindeling"),
    ("2021-2022", "Supersheet indeling 1.0 - 2021-2022.xlsx", "supersheet"),
    ("2022-2023", "Supersheet indeling 1.0 - 2022-2023.xlsx", "supersheet"),
    ("2023-2024", "Supersheet indeling 1.0 - 2023-2024.xlsx", "supersheet"),
    ("2024-2025", "Supersheet indeling 1.0 - 2024-2025.xlsx", "supersheet"),
    ("2025-2026", "Supersheet indeling 1.0 - 2025-2026.xlsx", "supersheet"),
]

# --- Filterpatronen ---

STAF_LABELS = re.compile(
    r"^(trainers?|coach\b|begeleiders?|verzorgers?|fysio|teambegeleiding|scheidsrechter)",
    re.IGNORECASE,
)
PLACEHOLDER_PATTERNS = re.compile(
    r"(selectie|vacature|trainerscarrousel|carrousel|\[.*\])",
    re.IGNORECASE,
)
GESTOPT_PATTERN = re.compile(r"\b(is gestopt|gestopt|stopt)\b", re.IGNORECASE)
SECTION_HEADER = re.compile(
    r"^(team\s+\d|[A-Z]{1,2}\d|S\d|MW\d|selectie|senioren|junioren|aspiranten|pupillen)",
    re.IGNORECASE,
)
EXCLUDE_SECTION = re.compile(r"(recreant|AR.?lijst|\bAR\b.*reserve)", re.IGNORECASE)

# Naam-achtig: minstens 2 chars, begint met letter, geen pure label
NAME_LIKE = re.compile(r"^[A-Za-zÀ-ÿ].{1,}", re.UNICODE)
NOT_A_NAME = re.compile(
    r"^(trainers?|coach\b|begeleiders?|verzorgers?|fysio|heren|dames|naam|staf|"
    r"team\b|selectie|senioren|junioren|aspiranten|pupillen|kangoeroes|"
    r"leeftijd|geb|roepnaam|rel|peildatum|notitie|blad|"
    r"opstelling|indeling|seizoen|\d{4})",
    re.IGNORECASE,
)


def normalize_name(name):
    """Normaliseer naam voor deduplicatie."""
    if not name:
        return ""
    s = str(name).strip()
    s = re.sub(r"^\d+\.?\s*", "", s)         # rugnummers
    s = re.sub(r"\s*\(.*?\)\s*", " ", s)     # annotaties
    s = re.sub(r"\s+", " ", s).strip()
    return s


def is_valid_name(value):
    """Is dit een geldige spelernaam?"""
    if not value:
        return False
    v = str(value).strip()
    if len(v) < 3:
        return False
    if NOT_A_NAME.match(v):
        return False
    if PLACEHOLDER_PATTERNS.search(v):
        return False
    if not NAME_LIKE.match(v):
        return False
    # Geen pure getallen
    if v.replace(".", "").replace(",", "").strip().isdigit():
        return False
    return True


# ============================================================
# Excel helpers
# ============================================================

def read_xls(filepath):
    """Lees .xls bestand."""
    import xlrd
    sheets = {}
    wb = xlrd.open_workbook(filepath)
    for name in wb.sheet_names():
        sheet = wb.sheet_by_name(name)
        rows = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                v = sheet.cell(r, c).value
                if isinstance(v, float) and v == int(v):
                    v = int(v)
                row.append(v if v != "" else None)
            rows.append(row)
        sheets[name] = rows
    return sheets


def read_xlsx(filepath):
    """Lees .xlsx/.xlsm bestand."""
    import openpyxl
    sheets = {}
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        sheets[name] = rows
    wb.close()
    return sheets


def read_excel(filepath):
    if filepath.endswith(".xls"):
        return read_xls(filepath)
    return read_xlsx(filepath)


# ============================================================
# Strategie B: Supersheets (2021-2026)
# ============================================================

def parse_supersheet(filepath, seizoen):
    """Lees Spelerslijst: twee panelen (V links, M rechts), gescheiden door lege kolom."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if "speler" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        return None

    rows = list(wb[sheet_name].iter_rows(values_only=True))
    wb.close()
    if not rows:
        return None

    # Vind header-rij
    header_idx = None
    for i, row in enumerate(rows):
        for cell in row:
            if cell and "rel" in str(cell).lower():
                header_idx = i
                break
        if header_idx is not None:
            break
    if header_idx is None:
        return None

    header = list(rows[header_idx])

    # Vind alle Rel.nr kolommen
    rel_cols = [j for j, h in enumerate(header) if h and "rel" in str(h).lower()]

    # Vind team-kolom per paneel
    szn_short = seizoen.split("-")[0][2:] + "-" + seizoen.split("-")[1][2:]

    def find_team_col(start, end):
        for j in range(start, min(end, len(header))):
            h = str(header[j] or "").lower()
            if "team" in h and ("peildatum" in h or szn_short in h or seizoen in h):
                return j
        # Fallback: eerste "team" kolom die niet "alt" of vorig seizoen is
        for j in range(start, min(end, len(header))):
            h = str(header[j] or "").lower()
            if h.startswith("team") and "alt" not in h:
                # Check dat het niet een vorig-seizoen kolom is
                prev_szn = str(int(seizoen[:4]) - 1)
                if prev_szn not in h:
                    return j
        return None

    panels = []
    if len(rel_cols) >= 2:
        tc_v = find_team_col(rel_cols[0] + 1, rel_cols[1])
        tc_m = find_team_col(rel_cols[1] + 1, len(header))
        if tc_v:
            panels.append((rel_cols[0], tc_v, "V"))
        if tc_m:
            panels.append((rel_cols[1], tc_m, "M"))
    elif len(rel_cols) == 1:
        tc = find_team_col(rel_cols[0] + 1, len(header))
        if tc:
            panels.append((rel_cols[0], tc, None))

    spelers = []
    uitgefilterd = {"stopt": 0, "geen_team": 0, "geen_relnr": 0}

    for rel_col, team_col, geslacht in panels:
        for row in rows[header_idx + 1:]:
            if rel_col >= len(row) or team_col >= len(row):
                continue

            rel_nr = row[rel_col]
            team = row[team_col]
            naam_col = rel_col + 1
            naam = str(row[naam_col]).strip() if naam_col < len(row) and row[naam_col] else ""

            if not rel_nr:
                # Kangoeroe zonder rel_nr
                if naam and team and str(team).strip().upper() == "K":
                    spelers.append({"naam": naam, "geslacht": geslacht, "team": "K", "rel_nr": None})
                    continue
                if naam:
                    uitgefilterd["geen_relnr"] += 1
                continue

            rel_nr = str(rel_nr).strip()
            team_str = str(team).strip() if team else ""

            if not team_str or team_str == "-":
                uitgefilterd["geen_team"] += 1
                continue
            if team_str.lower() in ("stopt", "gestopt", "stop"):
                uitgefilterd["stopt"] += 1
                continue

            spelers.append({"naam": naam, "geslacht": geslacht, "team": team_str, "rel_nr": rel_nr})

    # Dedup op rel_nr
    seen = set()
    unique = []
    for sp in spelers:
        key = sp["rel_nr"] or sp["naam"]
        if key not in seen:
            seen.add(key)
            unique.append(sp)

    heren = sum(1 for s in unique if s["geslacht"] == "M")
    dames = sum(1 for s in unique if s["geslacht"] == "V")

    return {
        "seizoen": seizoen,
        "totaal": len(unique),
        "heren": heren,
        "dames": dames,
        "onbekend": len(unique) - heren - dames,
        "teams": len(set(s["team"] for s in unique)),
        "uitgefilterd": uitgefilterd,
    }


# ============================================================
# Strategie A: Opstellingsbestanden (2010-2020)
# ============================================================

def determine_gender(rows, current_row, current_col):
    """Bepaal geslacht op basis van Heren/Dames headers erboven."""
    for scan_row in range(current_row - 1, max(current_row - 20, -1), -1):
        if scan_row < 0:
            continue
        row = rows[scan_row]
        heren_col = None
        dames_col = None
        for c, cell in enumerate(row):
            if cell and isinstance(cell, str):
                cl = cell.strip().lower()
                if cl in ("heren", "mannen", "jongens"):
                    heren_col = c
                elif cl in ("dames", "vrouwen", "meisjes"):
                    dames_col = c
        if heren_col is not None or dames_col is not None:
            if heren_col is not None and dames_col is not None:
                return "M" if abs(current_col - heren_col) <= abs(current_col - dames_col) else "V"
            return "M" if heren_col is not None else "V"
    return None


def parse_opstellingslijst(filepath, seizoen):
    """
    Parse opstellingslijsten. Twee varianten:
    - Variant 1 (2010-2012, 2016-2020): rugnummer (1-20) + naam in volgende cel
    - Variant 2 (2013-2015): namen direct in kolommen B (dames) en D (heren), geen rugnummers
    """
    sheets = read_excel(filepath)
    all_spelers = []
    uitgefilterd = {"trainers": 0, "placeholders": 0, "gestopt": 0, "recreanten_ar": 0}

    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        # Skip lege/admin sheets
        sn_lower = sheet_name.lower()
        if any(skip in sn_lower for skip in ("blad", "sheet", "check", "hulp", "controle")):
            continue

        in_excluded = False
        sheet_spelers_before = len(all_spelers)

        # Eerste pass: probeer rugnummer-variant
        for row_idx, row in enumerate(rows):
            row_text = " ".join(str(c) for c in row if c is not None)
            row_lower = row_text.lower()

            # Sectie-detectie
            if EXCLUDE_SECTION.search(row_lower):
                in_excluded = True
                continue
            if SECTION_HEADER.match(row_text.strip()):
                if not EXCLUDE_SECTION.search(row_lower):
                    in_excluded = False

            if in_excluded:
                uitgefilterd["recreanten_ar"] += 1
                continue

            for col_idx in range(len(row) - 1):
                cell = row[col_idx]
                next_cell = row[col_idx + 1] if col_idx + 1 < len(row) else None

                # Rugnummer: integer/float 1-20
                is_nr = False
                if isinstance(cell, (int, float)) and 1 <= cell <= 20:
                    is_nr = True
                elif isinstance(cell, str) and cell.strip().isdigit() and 1 <= int(cell.strip()) <= 20:
                    is_nr = True

                if not is_nr or not next_cell:
                    continue

                name = str(next_cell).strip()

                if STAF_LABELS.match(name):
                    uitgefilterd["trainers"] += 1
                    continue
                if PLACEHOLDER_PATTERNS.search(name):
                    uitgefilterd["placeholders"] += 1
                    continue

                # Check gestopt
                extra = str(row[col_idx + 2]).strip() if col_idx + 2 < len(row) and row[col_idx + 2] else ""
                if GESTOPT_PATTERN.search(name) or GESTOPT_PATTERN.search(extra):
                    uitgefilterd["gestopt"] += 1
                    continue

                if not is_valid_name(name):
                    continue

                geslacht = determine_gender(rows, row_idx, col_idx)

                # Fallback geslacht: positioneel (links=V, rechts=M)
                if geslacht is None:
                    nr_positions = set()
                    for scan_row in rows:
                        for sc in range(len(scan_row) - 1):
                            sv = scan_row[sc]
                            if isinstance(sv, (int, float)) and 1 <= sv <= 20:
                                snv = scan_row[sc + 1]
                                if snv and isinstance(snv, str) and len(str(snv).strip()) > 3:
                                    nr_positions.add(sc)
                    if len(nr_positions) >= 2:
                        sorted_pos = sorted(nr_positions)
                        midpoint = (sorted_pos[0] + sorted_pos[-1]) / 2
                        geslacht = "V" if col_idx < midpoint else "M"

                normalized = normalize_name(name)
                if normalized:
                    all_spelers.append({"naam": normalized, "geslacht": geslacht})

        found_by_number = len(all_spelers) > sheet_spelers_before

        # Als geen spelers gevonden via rugnummers in DEZE SHEET: variant 2
        # Format 2013-2016: c1=dames, c3=heren, c6/c7=staf
        if not found_by_number:
            in_excluded = False

            # Detecteer kolomstructuur: zoek kolommen met namen vs staf
            # In deze bestanden staan namen in specifieke kolommen (typisch c1=V, c3=M)
            # en staf in latere kolommen (c5+, c6+, c7+)
            player_cols = {}  # col_idx -> geslacht
            staf_cols = set()

            # Scan eerste 5 rijen voor structuur-hints
            for row_idx, row in enumerate(rows[:8]):
                for col_idx, cell in enumerate(row):
                    if cell and isinstance(cell, str):
                        cl = cell.strip().lower()
                        if cl in ("heren", "mannen", "jongens"):
                            player_cols[col_idx] = "M"
                        elif cl in ("dames", "vrouwen", "meisjes"):
                            player_cols[col_idx] = "V"
                        elif cl in ("staf", "trainers", "trainer", "coach"):
                            staf_cols.add(col_idx)

            # Als geen expliciete headers gevonden: gebruik positie-heuristiek
            # Typisch voor 2013-2016: c1=dames, c3=heren
            if not player_cols:
                # Tel hoeveel naam-achtige cellen per kolom
                col_name_counts = {}
                for row in rows:
                    for c, cell in enumerate(row):
                        if cell and isinstance(cell, str) and "," in cell and len(cell) > 5:
                            col_name_counts[c] = col_name_counts.get(c, 0) + 1
                # De twee kolommen met meeste namen zijn speler-kolommen
                top_cols = sorted(col_name_counts.items(), key=lambda x: -x[1])
                if len(top_cols) >= 2:
                    # Lagere kolom-index = dames (links), hogere = heren (rechts)
                    cols_sorted = sorted([top_cols[0][0], top_cols[1][0]])
                    player_cols[cols_sorted[0]] = "V"
                    player_cols[cols_sorted[1]] = "M"
                # Alle hogere kolommen zijn staf
                for c in col_name_counts:
                    if c not in player_cols and c >= 5:
                        staf_cols.add(c)

            for row_idx, row in enumerate(rows):
                row_text = " ".join(str(c) for c in row if c is not None)
                row_lower = row_text.lower()

                if EXCLUDE_SECTION.search(row_lower):
                    in_excluded = True
                    continue
                if SECTION_HEADER.match(row_text.strip()):
                    if not EXCLUDE_SECTION.search(row_lower):
                        in_excluded = False
                if in_excluded:
                    uitgefilterd["recreanten_ar"] += 1
                    continue

                for col_idx, cell in enumerate(row):
                    if not cell or not isinstance(cell, str):
                        continue
                    # Alleen speler-kolommen parsen (als we ze kennen)
                    if player_cols and col_idx not in player_cols:
                        if col_idx in staf_cols:
                            uitgefilterd["trainers"] += 1
                        continue
                    if col_idx in staf_cols:
                        uitgefilterd["trainers"] += 1
                        continue

                    name = cell.strip()

                    # Filter team-headers en notities die in naamkolommen staan
                    # Bijv. "S1/S2   4/4 +5/5", "Teamselectie na de zomer", "Geblesseerd"
                    if re.match(r"^[A-Z]\d.*\d/\d", name):  # team header met 4/4 formaat
                        continue
                    if re.match(r"^(In |Teamselectie|Geblesseerd|Let op|Opm|NB)", name, re.IGNORECASE):
                        continue

                    if not is_valid_name(name):
                        continue
                    if STAF_LABELS.match(name):
                        uitgefilterd["trainers"] += 1
                        continue
                    if PLACEHOLDER_PATTERNS.search(name):
                        uitgefilterd["placeholders"] += 1
                        continue
                    if GESTOPT_PATTERN.search(name):
                        uitgefilterd["gestopt"] += 1
                        continue

                    # Naam moet een komma bevatten (Achternaam, Voorletter) of meerdere woorden
                    if "," not in name and " " not in name:
                        continue

                    geslacht = player_cols.get(col_idx)
                    if not geslacht:
                        geslacht = determine_gender(rows, row_idx, col_idx)

                    normalized = normalize_name(name)
                    if normalized:
                        all_spelers.append({"naam": normalized, "geslacht": geslacht})

    # Dedup op naam
    seen = set()
    unique = []
    for sp in all_spelers:
        key = sp["naam"].lower()
        if key not in seen:
            seen.add(key)
            unique.append(sp)

    heren = sum(1 for s in unique if s["geslacht"] == "M")
    dames = sum(1 for s in unique if s["geslacht"] == "V")

    return {
        "seizoen": seizoen,
        "totaal": len(unique),
        "heren": heren,
        "dames": dames,
        "onbekend": len(unique) - heren - dames,
        "teams": None,
        "uitgefilterd": uitgefilterd,
    }


# ============================================================
# Strategie A2: A2 formulier (2018-2019)
# ============================================================

def parse_a2_formulier(filepath, seizoen):
    """Parse A2 formulier: alleen Leden sheet, filter op team-toewijzing."""
    sheets = read_excel(filepath)

    # Zoek het Leden sheet
    leden_rows = None
    for name, rows in sheets.items():
        if name.lower() == "leden":
            leden_rows = rows
            break
    if not leden_rows:
        return None

    # Zoek header-rij (bevat "Rel" of "Naam")
    header_idx = None
    for i, row in enumerate(leden_rows):
        for cell in row:
            if cell and "rel" in str(cell).lower():
                header_idx = i
                break
        if header_idx is not None:
            break
    if header_idx is None:
        return None

    header = leden_rows[header_idx]

    # Vind kolom-indices
    rel_col = None
    naam_col = None
    mv_col = None
    team_zaal_col = None
    team_veld_col = None

    for j, h in enumerate(header):
        hl = str(h or "").lower()
        if "rel" in hl and rel_col is None:
            rel_col = j
        if "naam" in hl and naam_col is None:
            naam_col = j
        if hl in ("mv", "m/v") or "geslacht" in hl:
            mv_col = j
        if "zaal" in hl and team_zaal_col is None:
            team_zaal_col = j
        if "veld" in hl and team_veld_col is None:
            team_veld_col = j

    if naam_col is None:
        return None

    spelers = []
    uitgefilterd = {"geen_team": 0, "stopt": 0}

    for row in leden_rows[header_idx + 1:]:
        if naam_col >= len(row) or not row[naam_col]:
            continue

        naam = str(row[naam_col]).strip()
        if not naam or len(naam) < 3:
            continue

        # Check of speler in een team zit (zaal of veld)
        has_team = False
        if team_zaal_col and team_zaal_col < len(row) and row[team_zaal_col]:
            t = str(row[team_zaal_col]).strip()
            if t and t.lower() not in ("x1", "x", "-", ""):
                has_team = True
        if team_veld_col and team_veld_col < len(row) and row[team_veld_col]:
            t = str(row[team_veld_col]).strip()
            if t and t.lower() not in ("x1", "x", "-", ""):
                has_team = True

        if not has_team:
            uitgefilterd["geen_team"] += 1
            continue

        geslacht = None
        if mv_col and mv_col < len(row) and row[mv_col]:
            g = str(row[mv_col]).strip().upper()
            if g in ("M", "V"):
                geslacht = g

        normalized = normalize_name(naam)
        if normalized:
            spelers.append({"naam": normalized, "geslacht": geslacht})

    # Dedup
    seen = set()
    unique = []
    for sp in spelers:
        key = sp["naam"].lower()
        if key not in seen:
            seen.add(key)
            unique.append(sp)

    heren = sum(1 for s in unique if s["geslacht"] == "M")
    dames = sum(1 for s in unique if s["geslacht"] == "V")

    return {
        "seizoen": seizoen,
        "totaal": len(unique),
        "heren": heren,
        "dames": dames,
        "onbekend": len(unique) - heren - dames,
        "teams": None,
        "uitgefilterd": uitgefilterd,
    }


# ============================================================
# Strategie A-teamindeling: Teamindeling 2020-2021
# ============================================================

def parse_teamindeling(filepath, seizoen):
    """Parse Teamindeling: namen onder Dames/Heren kolommen, Staf apart."""
    sheets = read_excel(filepath)
    all_spelers = []
    uitgefilterd = {"trainers": 0, "placeholders": 0, "gestopt": 0}

    skip_sheets = {"spelers", "actielijst", "gesprekje", "trainer voorzet"}

    for sheet_name, rows in sheets.items():
        if not rows:
            continue
        if any(skip in sheet_name.lower() for skip in skip_sheets):
            continue
        if "21-22" in sheet_name or "22-23" in sheet_name:
            continue

        # Vind gender- en staf-kolommen
        gender_cols = {}
        staf_cols = set()
        for row_idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if cell and isinstance(cell, str):
                    cl = cell.strip().lower()
                    if cl in ("heren", "mannen", "jongens"):
                        gender_cols[col_idx] = "M"
                    elif cl in ("dames", "vrouwen", "meisjes"):
                        gender_cols[col_idx] = "V"
                    elif cl == "staf":
                        staf_cols.add(col_idx)

        for row_idx, row in enumerate(rows):
            for col_idx, cell in enumerate(row):
                if not cell or not isinstance(cell, str):
                    continue
                if col_idx in staf_cols:
                    # Scan boven: check of "Staf" header van toepassing is
                    is_staf = False
                    for sr in range(row_idx - 1, max(row_idx - 15, -1), -1):
                        if sr < 0:
                            break
                        r = rows[sr]
                        if col_idx < len(r) and r[col_idx]:
                            ht = str(r[col_idx]).strip().lower()
                            if ht == "staf":
                                is_staf = True
                            break
                    if is_staf:
                        uitgefilterd["trainers"] += 1
                        continue

                name = cell.strip()
                if not is_valid_name(name):
                    continue
                if STAF_LABELS.match(name):
                    uitgefilterd["trainers"] += 1
                    continue
                if PLACEHOLDER_PATTERNS.search(name):
                    uitgefilterd["placeholders"] += 1
                    continue

                geslacht = gender_cols.get(col_idx) or determine_gender(rows, row_idx, col_idx)
                normalized = normalize_name(name)
                if normalized:
                    all_spelers.append({"naam": normalized, "geslacht": geslacht})

    seen = set()
    unique = []
    for sp in all_spelers:
        key = sp["naam"].lower()
        if key not in seen:
            seen.add(key)
            unique.append(sp)

    heren = sum(1 for s in unique if s["geslacht"] == "M")
    dames = sum(1 for s in unique if s["geslacht"] == "V")

    return {
        "seizoen": seizoen,
        "totaal": len(unique),
        "heren": heren,
        "dames": dames,
        "onbekend": len(unique) - heren - dames,
        "teams": None,
        "uitgefilterd": uitgefilterd,
    }


# ============================================================
# Telling-bestand: autoritatieve bron 2010-2019
# ============================================================

TELLING_SEIZOENEN = [
    "2010-2011", "2011-2012", "2012-2013", "2013-2014",
    "2014-2015", "2015-2016", "2016-2017", "2017-2018",
    "2019-2020",  # 2018-2019 ontbreekt in Telling
]


def parse_telling(filepath):
    """
    Parse het hand-gecureerde Telling-bestand.
    Leest:
    - Samenvatting-sheet: totalen H/D/Totaal per seizoen (validatie)
    - Per seizoenssheet: individuele spelers met Naam, H/D, Team
    - Details per team-sheet: per-team uitsplitsing per seizoen
    Retourneert dict met per seizoen: totalen + spelerslijsten + per-team details.
    """
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    telling_data = {}

    # --- Stap 1: Samenvatting lezen (referentiecijfers) ---
    samenvatting = {}
    for name in wb.sheetnames:
        if "samenvatting" in name.lower():
            rows = list(wb[name].iter_rows(values_only=True))
            # Zoek header-rij met seizoenskolommen
            for i, row in enumerate(rows):
                for j, cell in enumerate(row):
                    if cell and "2010" in str(cell):
                        # Dit is de rij met seizoenslabels
                        seizoen_cols = {}
                        for k in range(j, len(row)):
                            val = str(row[k] or "").strip()
                            if re.match(r"20\d{2}-20\d{2}", val):
                                seizoen_cols[k] = val
                        # Lees H/D/Totaal uit de volgende rijen
                        for data_row in rows[i+1:i+10]:
                            label = str(data_row[j-1] if j > 0 else data_row[0] or "").strip().lower()
                            for k, szn in seizoen_cols.items():
                                if szn not in samenvatting:
                                    samenvatting[szn] = {}
                                if k < len(data_row) and data_row[k] is not None:
                                    try:
                                        val = int(data_row[k])
                                    except (ValueError, TypeError):
                                        continue
                                    if "heren" in label or label == "h":
                                        samenvatting[szn]["heren"] = val
                                    elif "dames" in label or label == "d":
                                        samenvatting[szn]["dames"] = val
                                    elif "totaal" in label:
                                        samenvatting[szn]["totaal"] = val
                        break
                if samenvatting:
                    break
            break

    # --- Stap 2: Details per team lezen ---
    team_details = {}
    for name in wb.sheetnames:
        if "details" in name.lower() or "per team" in name.lower():
            rows = list(wb[name].iter_rows(values_only=True))
            # Zoek seizoenskolommen
            for i, row in enumerate(rows):
                for j, cell in enumerate(row):
                    if cell and "2010" in str(cell):
                        seizoen_cols = {}
                        for k in range(j, len(row)):
                            val = str(row[k] or "").strip()
                            if re.match(r"20\d{2}-20\d{2}", val):
                                seizoen_cols[k] = val
                        # Lees team-rijen
                        for data_row in rows[i+1:]:
                            team_label = str(data_row[0] or "").strip() if data_row[0] else ""
                            if not team_label or team_label.lower() in ("totaal", ""):
                                if team_label.lower() == "totaal":
                                    continue
                                if not team_label:
                                    continue
                            for k, szn in seizoen_cols.items():
                                if szn not in team_details:
                                    team_details[szn] = {}
                                if k < len(data_row) and data_row[k] is not None:
                                    try:
                                        val = int(data_row[k])
                                    except (ValueError, TypeError):
                                        continue
                                    if val > 0:
                                        team_details[szn][team_label] = val
                        break
                if team_details:
                    break
            break

    # --- Stap 3: Individuele seizoenssheets lezen ---
    for name in wb.sheetnames:
        szn_match = re.match(r"(20\d{2})[-\s]*(20\d{2})", name)
        if not szn_match:
            continue
        seizoen = f"{szn_match.group(1)}-{szn_match.group(2)}"
        if seizoen not in TELLING_SEIZOENEN:
            continue

        rows = list(wb[name].iter_rows(values_only=True))
        if not rows:
            continue

        # Zoek header-rij
        header_idx = None
        naam_col = None
        hd_col = None
        team_col = None
        for i, row in enumerate(rows):
            for j, cell in enumerate(row):
                val = str(cell or "").strip().lower()
                if val in ("naam", "speler"):
                    header_idx = i
                    naam_col = j
                    break
            if header_idx is not None:
                break

        if header_idx is None:
            # Probeer eerste rij als header
            header_idx = 0
            for j, cell in enumerate(rows[0]):
                val = str(cell or "").strip().lower()
                if "naam" in val:
                    naam_col = j
                elif val in ("h/d", "m/v", "geslacht", "hd"):
                    hd_col = j
                elif "team" in val:
                    team_col = j
            if naam_col is None:
                continue

        # Vind H/D en Team kolommen
        header_row = rows[header_idx]
        for j, cell in enumerate(header_row):
            val = str(cell or "").strip().lower()
            if val in ("h/d", "m/v", "geslacht", "hd") and hd_col is None:
                hd_col = j
            elif "team" in val and team_col is None:
                team_col = j

        spelers = []
        for row in rows[header_idx + 1:]:
            if naam_col >= len(row) or not row[naam_col]:
                continue
            naam = str(row[naam_col]).strip()
            if not naam or len(naam) < 2:
                continue

            geslacht = None
            if hd_col is not None and hd_col < len(row) and row[hd_col]:
                g = str(row[hd_col]).strip().upper()
                if g in ("H", "M"):
                    geslacht = "M"
                elif g in ("D", "V"):
                    geslacht = "V"

            team = None
            if team_col is not None and team_col < len(row) and row[team_col]:
                team = str(row[team_col]).strip()

            spelers.append({"naam": naam, "geslacht": geslacht, "team": team})

        heren = sum(1 for s in spelers if s["geslacht"] == "M")
        dames = sum(1 for s in spelers if s["geslacht"] == "V")

        # Per-team uitsplitsing uit spelerslijst
        teams_uit_spelers = {}
        for sp in spelers:
            t = sp["team"] or "Onbekend"
            if t not in teams_uit_spelers:
                teams_uit_spelers[t] = {"heren": [], "dames": []}
            if sp["geslacht"] == "M":
                teams_uit_spelers[t]["heren"].append(sp["naam"])
            elif sp["geslacht"] == "V":
                teams_uit_spelers[t]["dames"].append(sp["naam"])

        telling_data[seizoen] = {
            "totaal": len(spelers),
            "heren": heren,
            "dames": dames,
            "spelers": spelers,
            "teams": teams_uit_spelers,
            "samenvatting": samenvatting.get(seizoen, {}),
            "team_details": team_details.get(seizoen, {}),
        }

    wb.close()
    return telling_data


# ============================================================
# Main
# ============================================================

def main():
    results = []

    # --- Stap 1: Parse Telling-bestand (autoritatief voor 2010-2019) ---
    telling_path = os.path.join(TEAMINDELINGEN_DIR, TELLING_BESTAND)
    telling_data = {}
    if os.path.exists(telling_path):
        print(f"\n{'='*60}")
        print(f"  Telling-bestand laden: {TELLING_BESTAND}")
        print(f"{'='*60}")
        try:
            telling_data = parse_telling(telling_path)
            print(f"  Seizoenen gevonden: {', '.join(sorted(telling_data.keys()))}")
            for szn in sorted(telling_data.keys()):
                td = telling_data[szn]
                sam = td.get("samenvatting", {})
                check = ""
                if sam.get("totaal") and sam["totaal"] != td["totaal"]:
                    check = f"  ! samenvatting={sam['totaal']}"
                print(f"    {szn}: {td['totaal']} ({td['heren']}H/{td['dames']}D){check}")
        except Exception as e:
            print(f"  FOUT bij parsen Telling: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"WAARSCHUWING: Telling-bestand niet gevonden: {telling_path}")

    # --- Stap 2: Per seizoen de juiste bron kiezen ---
    for seizoen, bestand, strategie in BESTANDEN:
        # Gebruik Telling als primaire bron (als beschikbaar)
        if seizoen in telling_data:
            td = telling_data[seizoen]
            result = {
                "seizoen": seizoen,
                "totaal": td["totaal"],
                "heren": td["heren"],
                "dames": td["dames"],
                "onbekend": td["totaal"] - td["heren"] - td["dames"],
                "teams": len(td["teams"]),
                "bestand": TELLING_BESTAND,
                "strategie": "telling",
            }
            results.append(result)
            print(f"\n  {seizoen}: Telling -> {result['totaal']} ({result['heren']}H/{result['dames']}D)")
            continue

        # Fallback: originele parser
        filepath = os.path.join(TEAMINDELINGEN_DIR, bestand)
        if not os.path.exists(filepath):
            print(f"SKIP {seizoen}: bestand niet gevonden ({bestand})")
            results.append({"seizoen": seizoen, "bestand": bestand, "fout": "bestand niet gevonden"})
            continue

        print(f"\n{'='*60}")
        print(f"  {seizoen} -- {bestand} ({strategie})")
        print(f"{'='*60}")

        try:
            if strategie == "supersheet":
                result = parse_supersheet(filepath, seizoen)
            elif strategie == "opstellingslijst":
                result = parse_opstellingslijst(filepath, seizoen)
            elif strategie == "teamindeling":
                result = parse_teamindeling(filepath, seizoen)
            elif strategie == "a2_formulier":
                result = parse_a2_formulier(filepath, seizoen)
            else:
                print(f"  Onbekende strategie: {strategie}")
                continue

            if result:
                result["bestand"] = bestand
                result["strategie"] = strategie
                results.append(result)
                print(f"  Totaal: {result['totaal']}")
                print(f"  Heren:  {result['heren']}  Dames: {result['dames']}  Onbekend: {result.get('onbekend', 0)}")
                if result.get("teams"):
                    print(f"  Teams:  {result['teams']}")
                if result.get("uitgefilterd"):
                    print(f"  Uitgefilterd: {result['uitgefilterd']}")
            else:
                print(f"  FOUT: kon niet parsen")
                results.append({"seizoen": seizoen, "bestand": bestand, "fout": "kon niet parsen"})

        except Exception as e:
            print(f"  FOUT: {e}")
            import traceback
            traceback.print_exc()
            results.append({"seizoen": seizoen, "bestand": bestand, "fout": str(e)})

    # --- Samenvatting ---
    print(f"\n{'='*60}")
    print("  SAMENVATTING")
    print(f"{'='*60}")
    print(f"{'Seizoen':<14} {'Totaal':>7} {'Heren':>7} {'Dames':>7} {'Onbek':>7}  {'Strategie':<18}")
    print("-" * 68)
    for r in results:
        if "fout" in r:
            print(f"{r['seizoen']:<14} {'FOUT':>7}   {r.get('fout', '')}")
        else:
            print(f"{r['seizoen']:<14} {r['totaal']:>7} {r['heren']:>7} {r['dames']:>7} {r.get('onbekend', 0):>7}  {r.get('strategie', ''):<18}")

    # --- Output 1: actieve-spelers-per-seizoen.json ---
    output = {
        "_meta": {
            "gegenereerd": date.today().isoformat(),
            "bron": "bronbestanden docs/teamindelingen/",
            "beschrijving": "Unieke ingedeelde korfbalspelers per seizoen (competitie + midweek, excl. recreanten/AR/staf)",
        },
        "seizoenen": results,
    }

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)
    print(f"\nOutput: {OUTPUT_PATH}")

    # --- Output 2: teamindelingen-telling.json (gedetailleerde Telling-data) ---
    if telling_data:
        telling_output = {
            "_meta": {
                "gegenereerd": date.today().isoformat(),
                "bron": TELLING_BESTAND,
                "beschrijving": "Gedetailleerde teamindelingen uit hand-gecureerd Telling-bestand (2010-2019)",
            },
            "seizoenen": {},
        }
        for szn in sorted(telling_data.keys()):
            td = telling_data[szn]
            telling_output["seizoenen"][szn] = {
                "totaal": td["totaal"],
                "heren": td["heren"],
                "dames": td["dames"],
                "teams": td["teams"],
                "spelers": td["spelers"],
                "samenvatting_check": td.get("samenvatting", {}),
                "team_details": td.get("team_details", {}),
            }

        os.makedirs(os.path.dirname(TELLING_OUTPUT_PATH), exist_ok=True)
        with open(TELLING_OUTPUT_PATH, "w", encoding="utf-8") as f:
            json.dump(telling_output, f, ensure_ascii=False, indent=2)
        print(f"Output: {TELLING_OUTPUT_PATH}")

    # --- Output 3: hertelling.json updaten met Telling-cijfers ---
    if telling_data and os.path.exists(HERTELLING_PATH):
        with open(HERTELLING_PATH, "r", encoding="utf-8") as f:
            hertelling = json.load(f)

        for szn, td in telling_data.items():
            # Vervang de spelerspaden-afgeleide data met Telling-data
            teams_mv = {}
            for team_naam, team_data in td["teams"].items():
                teams_mv[team_naam] = {
                    "M": len(team_data.get("heren", [])),
                    "V": len(team_data.get("dames", [])),
                }
            hertelling[szn] = {
                "totaal": td["totaal"],
                "M": td["heren"],
                "V": td["dames"],
                "teams": teams_mv,
                "aantal_teams": len(td["teams"]),
                "bron": "telling",
            }

        with open(HERTELLING_PATH, "w", encoding="utf-8") as f:
            json.dump(hertelling, f, ensure_ascii=False, indent=2)
        print(f"Output: {HERTELLING_PATH} (updated)")


if __name__ == "__main__":
    main()
