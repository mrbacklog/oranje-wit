"""
Extract spelerdata uit een Supersheet Excel-bestand en genereer:
- Verrijkte snapshot (data/leden/snapshots/YYYY-06-01.json)
- Aggregaties per geboortejaar, team en kleur (data/aggregaties/)

Gebruik:
  python scripts/extract-supersheet.py "docs/teamindelingen/Supersheet indeling 1.0 - 2024-2025.xlsx" 2024-2025
"""

import sys
import json
import re
from pathlib import Path
from datetime import datetime
from collections import defaultdict

import openpyxl

ROOT = Path(__file__).resolve().parent.parent.parent
MAPPING_FILE = ROOT / "data" / "modellen" / "categorie-mapping.json"


def load_categorie_mapping():
    with open(MAPPING_FILE, encoding="utf-8") as f:
        data = json.load(f)
    return data["categorie_naar_band"]


def detect_sections(ws):
    """Detecteer dames- en heren-kolommen door de header-rij te scannen."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(1, col).value
        if val:
            headers[col] = str(val).strip()

    # Zoek de scheidingskolom: eerste kolom zonder header na de dames-sectie
    dames_cols = {}
    heren_cols = {}
    found_gap = False
    last_dames_col = 0

    for col in range(1, ws.max_column + 1):
        if col in headers:
            if not found_gap:
                dames_cols[col] = headers[col]
                last_dames_col = col
            else:
                heren_cols[col] = headers[col]
        elif not found_gap and last_dames_col > 0:
            found_gap = True

    return dames_cols, heren_cols


def find_column(cols, *keywords):
    """Zoek een kolom op basis van keywords in de header."""
    for col, header in cols.items():
        header_lower = header.lower()
        for kw in keywords:
            if kw.lower() in header_lower:
                return col
    return None


def find_team_column(cols, seizoen):
    """Zoek de teamkolom voor het huidige seizoen.

    Prioriteit:
    1. "Team op Peildatum" — de primaire teamtoewijzing in alle Supersheets
    2. Exacte seizoensmatch (bv. "Team 25-26") — voor 2025-2026 Supersheet
    """
    # Prioriteit 1: "Team op Peildatum"
    for col, header in cols.items():
        h = header.lower()
        if "team" in h and "peildatum" in h:
            return col

    # Prioriteit 2: exacte seizoensmatch
    jaar1, jaar2 = seizoen.split("-")
    kort = f"{jaar1[2:]}-{jaar2[2:]}"  # bv. "25-26"
    for col, header in cols.items():
        h = header.lower().replace(" ", "")
        if kort in h.replace(" ", ""):
            if "alt" not in h:
                return col

    return None


def parse_naam(naam_str):
    """Parse 'Groot, S. (Senna) de' → (roepnaam, achternaam, tussenvoegsel)."""
    if not naam_str or not isinstance(naam_str, str):
        return None, None, None

    naam_str = naam_str.strip()

    # Zoek roepnaam tussen haakjes
    roepnaam = None
    match = re.search(r"\(([^)]+)\)", naam_str)
    if match:
        roepnaam = match.group(1).strip()

    # Splits op komma: "Groot, S. (Senna) de" → ["Groot", " S. (Senna) de"]
    parts = naam_str.split(",", 1)
    if len(parts) == 2:
        achternaam = parts[0].strip()
        rest = parts[1].strip()
        # Verwijder initialen en roepnaam uit rest voor tussenvoegsel
        rest = re.sub(r"\([^)]*\)", "", rest)  # verwijder (roepnaam)
        rest = re.sub(r"[A-Z]\.", "", rest)  # verwijder initialen
        tussenvoegsel = rest.strip() or None
    else:
        achternaam = naam_str
        tussenvoegsel = None

    return roepnaam, achternaam, tussenvoegsel


def extract_section(ws, cols, geslacht, team_col, max_row):
    """Extraheer spelers uit een sectie (dames of heren)."""
    rel_col = find_column(cols, "rel")
    naam_col = find_column(cols, "naam")
    roepnaam_col = find_column(cols, "roepnaam")
    geb_col = find_column(cols, "geb")
    leeftijd_col = find_column(cols, "leeftijd")

    if not rel_col:
        return []

    spelers = []
    for row in range(2, max_row + 1):
        rel_code = ws.cell(row, rel_col).value
        if not rel_code:
            continue
        rel_code = str(rel_code).strip()

        # Filter sectie-markers en ongeldige rel_codes
        if rel_code.startswith("-") or rel_code.startswith("XXX"):
            continue
        # Rel_codes zijn alfanumeriek, typisch 7 karakters
        if len(rel_code) < 5 or not any(c.isalpha() for c in rel_code):
            continue

        # Naam parsing
        naam_raw = ws.cell(row, naam_col).value if naam_col else None
        roepnaam_raw = ws.cell(row, roepnaam_col).value if roepnaam_col else None
        roepnaam_parsed, achternaam, tussenvoegsel = parse_naam(naam_raw)
        roepnaam = str(roepnaam_raw).strip() if roepnaam_raw else roepnaam_parsed

        # Geboortejaar
        geb_datum = ws.cell(row, geb_col).value if geb_col else None
        geboortejaar = None
        if isinstance(geb_datum, datetime):
            geboortejaar = geb_datum.year
        elif isinstance(geb_datum, (int, float)) and geb_datum > 1900:
            geboortejaar = int(geb_datum)

        # Team
        team = None
        if team_col:
            team_val = ws.cell(row, team_col).value
            team_str = str(team_val).strip()
            if team_val and team_str not in ("-", "", "Stopt", "stopt", "?", "AR"):
                team = team_str

        spelers.append({
            "rel_code": rel_code,
            "roepnaam": roepnaam,
            "achternaam": achternaam,
            "tussenvoegsel": tussenvoegsel,
            "geslacht": geslacht,
            "geboortejaar": geboortejaar,
            "team": team,
        })

    return spelers


def assign_kleur(team, geboortejaar, peildatum_year, mapping):
    """Wijs kleur/band toe op basis van teamletter of leeftijd."""
    if not team:
        return None, "b"

    def kleur_uit_leeftijd():
        """Bepaal kleur/band op basis van leeftijd en mapping."""
        if not geboortejaar:
            return None, "b"
        leeftijd = peildatum_year - geboortejaar
        for cat, info in mapping.items():
            if info["leeftijden"] and leeftijd in info["leeftijden"]:
                return info["band"], "b"
        if leeftijd >= 19:
            return "Senioren", "a"
        return None, "b"

    # Nieuw systeem: "OW J1", "OW J17" etc. (jeugd B-categorie)
    if re.match(r"^(OW\s+)?J\d+", team):
        return kleur_uit_leeftijd()

    # A-categorie jeugd: "U15-1", "U17", "U19-2" etc.
    if re.match(r"^U\d+", team):
        return kleur_uit_leeftijd()

    # Senioren nieuw systeem: "S1S2", "S3", "S4", "S5", "S6"
    if re.match(r"^S\d", team):
        return "Senioren", "a"

    # Senioren: "MW1" (midweek)
    if team.startswith("MW"):
        return "Senioren", "a"

    # Senioren: puur numeriek team ("1", "2", "3")
    if re.match(r"^\d+$", team):
        return "Senioren", "a"

    # Kangoeroe
    if team.upper() == "K":
        return "Kangoeroe", "b"

    # Oud systeem: "Sen" (2021-2022)
    if team in ("Sen", "sen"):
        return "Senioren", "a"

    # Oud systeem: "A3", "B1", "C", "D", "E2", "F", "S5-6-7"
    letter = team[0].upper()
    if letter in mapping:
        info = mapping[letter]
        if letter == "S":
            return "Senioren", "a"
        return info["band"], "b"

    return None, "b"


def load_niet_spelend_set():
    """Laad rel_codes van niet-spelende leden uit de Sportlink snapshot.

    Gebruikt om niet-spelende leden die in het oude systeem als 'S'/'Sen'
    op de Spelerslijst stonden te filteren.
    """
    snapshots_dir = ROOT / "data" / "leden" / "snapshots"
    niet_spelend = set()
    for path in sorted(snapshots_dir.glob("*.json"), reverse=True):
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        # Zoek een Sportlink-snapshot (heeft geen bron_type of bron_type != supersheet)
        if data["_meta"].get("bron_type") == "supersheet":
            continue
        for lid in data["leden"]:
            if lid.get("spelactiviteit") in ("niet-spelend", "biljart", "kangoeroe-klup", "algemeen-reserve"):
                niet_spelend.add(lid["rel_code"])
        break  # Gebruik de eerste (nieuwste) Sportlink snapshot
    return niet_spelend


def filter_niet_spelend_senioren(spelers, niet_spelend_set):
    """Filter niet-spelende leden met senioren-team uit de spelerslijst.

    In het oude systeem (2024-2025 en eerder) stonden bestuursleden, ouders
    en oud-spelers op de Spelerslijst met team 'S' of 'Sen'. Deze mensen
    zijn geen actieve korfbalspelers.
    """
    if not niet_spelend_set:
        return spelers, 0

    senioren_teams = {"S", "Sen"}
    gefilterd = []
    verwijderd = 0
    for sp in spelers:
        team = sp.get("team")
        if team in senioren_teams and sp["rel_code"] in niet_spelend_set:
            verwijderd += 1
            continue
        gefilterd.append(sp)
    return gefilterd, verwijderd


def enrich_with_leden_tab(ws_leden, spelers):
    """Verrijk spelers met geslacht uit de leden-tab als die bestaat."""
    if not ws_leden:
        return

    # Bouw lookup van rel_code → geslacht
    leden_lookup = {}
    for row in range(2, ws_leden.max_row + 1):
        rel = ws_leden.cell(row, 1).value
        if not rel:
            continue
        geslacht = None
        # Zoek geslacht-kolom (typisch kolom 7)
        for col in range(1, ws_leden.max_column + 1):
            header = ws_leden.cell(1, col).value
            if header and "geslacht" in str(header).lower():
                geslacht_val = ws_leden.cell(row, col).value
                if geslacht_val:
                    geslacht_str = str(geslacht_val).strip().upper()
                    if geslacht_str in ("M", "MAN"):
                        geslacht = "M"
                    elif geslacht_str in ("V", "VROUW"):
                        geslacht = "V"
                break
        if geslacht:
            leden_lookup[str(rel).strip()] = geslacht

    # Verrijk
    for speler in spelers:
        if speler["rel_code"] in leden_lookup:
            speler["geslacht"] = leden_lookup[speler["rel_code"]]


def build_snapshot(spelers, seizoen, peildatum, bronbestand, mapping):
    """Bouw de snapshot JSON-structuur."""
    peildatum_year = int(peildatum.split("-")[0])

    leden = []
    for sp in spelers:
        kleur, categorie = assign_kleur(
            sp["team"], sp["geboortejaar"], peildatum_year, mapping
        )
        leeftijd = None
        if sp["geboortejaar"]:
            leeftijd = peildatum_year - sp["geboortejaar"]

        leden.append({
            "rel_code": sp["rel_code"],
            "roepnaam": sp["roepnaam"],
            "achternaam": sp["achternaam"],
            "tussenvoegsel": sp["tussenvoegsel"],
            "geslacht": sp["geslacht"],
            "geboortejaar": sp["geboortejaar"],
            "lidsoort": None,
            "spelactiviteit": "korfbal",
            "lid_sinds": None,
            "status": "actief",
            "team": sp["team"],
            "teamrol": None,
            "categorie": categorie,
            "kleur": kleur,
            "pool_veld": None,
            "pool_zaal": None,
            "a_categorie": None,
            "a_jaars": None,
            "leeftijd_peildatum": leeftijd,
        })

    # Sorteer op achternaam
    leden.sort(key=lambda x: (x["achternaam"] or "", x["roepnaam"] or ""))

    return {
        "_meta": {
            "snapshot_datum": peildatum,
            "seizoen": seizoen,
            "bronnen": [bronbestand],
            "totaal_leden": len(leden),
            "totaal_spelers": len([l for l in leden if l["team"]]),
            "bron_type": "supersheet",
        },
        "leden": leden,
    }


def build_aggregatie_geboortejaar(snapshot):
    """Genereer per-geboortejaar aggregatie."""
    counts = defaultdict(lambda: {"M": 0, "V": 0})
    for lid in snapshot["leden"]:
        if lid["geboortejaar"] and lid["team"]:
            key = (lid["geboortejaar"], lid["geslacht"])
            counts[key]["count"] = counts.get(key, {}).get("count", 0) + 1

    # Herstructureer
    data = []
    teller = defaultdict(int)
    for lid in snapshot["leden"]:
        if lid["geboortejaar"] and lid["team"]:
            key = (lid["geboortejaar"], lid["geslacht"])
            teller[key] += 1

    for (geboortejaar, geslacht), aantal in sorted(teller.items()):
        data.append({
            "geboortejaar": geboortejaar,
            "geslacht": geslacht,
            "aantal": aantal,
            "a_categorie": None,
            "a_jaars": None,
            "streef": None,
            "vulgraad": None,
            "signalering": None,
        })

    return {
        "_meta": {
            "datum": snapshot["_meta"]["snapshot_datum"],
            "seizoen": snapshot["_meta"]["seizoen"],
        },
        "data": data,
    }


def build_aggregatie_team(snapshot):
    """Genereer per-team aggregatie."""
    teams = defaultdict(lambda: {"M": 0, "V": 0, "leeftijden": []})
    for lid in snapshot["leden"]:
        if lid["team"]:
            t = teams[lid["team"]]
            t[lid["geslacht"]] += 1
            if lid["leeftijd_peildatum"]:
                t["leeftijden"].append(lid["leeftijd_peildatum"])

    # Zoek categorie en kleur per team
    team_meta = {}
    for lid in snapshot["leden"]:
        if lid["team"] and lid["team"] not in team_meta:
            team_meta[lid["team"]] = {
                "categorie": lid["categorie"],
                "kleur": lid["kleur"],
            }

    data = []
    for team_naam in sorted(teams.keys()):
        t = teams[team_naam]
        meta = team_meta.get(team_naam, {})
        gem = round(sum(t["leeftijden"]) / len(t["leeftijden"]), 1) if t["leeftijden"] else None
        data.append({
            "team": team_naam,
            "categorie": meta.get("categorie", "b"),
            "kleur": meta.get("kleur"),
            "niveau": None,
            "spelers_M": t["M"],
            "spelers_V": t["V"],
            "totaal": t["M"] + t["V"],
            "gem_leeftijd": gem,
        })

    return {
        "_meta": {
            "datum": snapshot["_meta"]["snapshot_datum"],
            "seizoen": snapshot["_meta"]["seizoen"],
        },
        "data": data,
    }


def build_aggregatie_kleur(snapshot):
    """Genereer per-kleur aggregatie."""
    kleuren = defaultdict(lambda: {"teams": set(), "M": 0, "V": 0})
    for lid in snapshot["leden"]:
        if lid["team"] and lid["kleur"]:
            k = kleuren[(lid["kleur"], lid["categorie"])]
            k["teams"].add(lid["team"])
            k[lid["geslacht"]] += 1

    data = []
    for (kleur, categorie) in sorted(kleuren.keys()):
        k = kleuren[(kleur, categorie)]
        data.append({
            "kleur": kleur,
            "categorie": categorie,
            "teams": len(k["teams"]),
            "spelers_M": k["M"],
            "spelers_V": k["V"],
            "totaal": k["M"] + k["V"],
        })

    return {
        "_meta": {
            "datum": snapshot["_meta"]["snapshot_datum"],
            "seizoen": snapshot["_meta"]["seizoen"],
        },
        "data": data,
    }


def write_json(data, path):
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  Geschreven: {path.relative_to(ROOT)}")


def main():
    if len(sys.argv) < 3:
        print("Gebruik: python scripts/extract-supersheet.py <excel-pad> <seizoen>")
        print("Voorbeeld: python scripts/extract-supersheet.py \"docs/teamindelingen/Supersheet indeling 1.0 - 2024-2025.xlsx\" 2024-2025")
        sys.exit(1)

    excel_path = ROOT / sys.argv[1]
    seizoen = sys.argv[2]
    jaar1 = seizoen.split("-")[0]
    peildatum = f"{jaar1}-06-01"

    print(f"Extractie: {excel_path.name}")
    print(f"Seizoen:   {seizoen}")
    print(f"Peildatum: {peildatum}")
    print()

    # Laad categorie-mapping
    mapping = load_categorie_mapping()

    # Open workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb["Spelerslijst"]

    # Detecteer structuur
    dames_cols, heren_cols = detect_sections(ws)
    print(f"Dames kolommen: {list(dames_cols.values())}")
    print(f"Heren kolommen: {list(heren_cols.values())}")

    # Zoek team-kolom per sectie
    dames_team_col = find_team_column(dames_cols, seizoen)
    heren_team_col = find_team_column(heren_cols, seizoen)
    print(f"Dames team-kolom: {dames_team_col} ({dames_cols.get(dames_team_col, '?')})")
    print(f"Heren team-kolom: {heren_team_col} ({heren_cols.get(heren_team_col, '?')})")
    print()

    # Extraheer spelers
    dames = extract_section(ws, dames_cols, "V", dames_team_col, ws.max_row)
    heren = extract_section(ws, heren_cols, "M", heren_team_col, ws.max_row)
    print(f"Geëxtraheerd: {len(dames)} dames, {len(heren)} heren")

    # Verrijk met leden-tab indien aanwezig
    if "leden" in wb.sheetnames:
        print("Leden-tab gevonden, verrijking geslacht...")
        enrich_with_leden_tab(wb["leden"], dames + heren)

    alle_spelers = dames + heren

    # Filter niet-spelende senioren (bestuursleden, ouders, oud-spelers)
    niet_spelend = load_niet_spelend_set()
    if niet_spelend:
        alle_spelers, verwijderd = filter_niet_spelend_senioren(alle_spelers, niet_spelend)
        if verwijderd:
            print(f"Niet-spelende senioren gefilterd: {verwijderd}")

    # Bouw snapshot
    bronbestand = f"docs/teamindelingen/{excel_path.name}"
    snapshot = build_snapshot(alle_spelers, seizoen, peildatum, bronbestand, mapping)
    print(f"Snapshot: {snapshot['_meta']['totaal_leden']} leden, {snapshot['_meta']['totaal_spelers']} met team")
    print()

    # Schrijf bestanden
    snapshot_path = ROOT / "data" / "leden" / "snapshots" / f"{peildatum}.json"
    write_json(snapshot, snapshot_path)

    agg_gj = build_aggregatie_geboortejaar(snapshot)
    write_json(agg_gj, ROOT / "data" / "aggregaties" / f"{peildatum}-per-geboortejaar.json")

    agg_team = build_aggregatie_team(snapshot)
    write_json(agg_team, ROOT / "data" / "aggregaties" / f"{peildatum}-per-team.json")

    agg_kleur = build_aggregatie_kleur(snapshot)
    write_json(agg_kleur, ROOT / "data" / "aggregaties" / f"{peildatum}-per-kleur.json")

    print()
    print("Klaar!")


if __name__ == "__main__":
    main()
