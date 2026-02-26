"""
Parse alle teamindelingsbestanden (2010-2026) naar een gestructureerde spelerspaden dataset.

Bronnen:
- Supersheets 2021-2026: Rel.nr, naam, roepnaam, geboortedatum, team, vorig team, stoppers
- A2 formulieren 2018-2022: Rel.nr, naam, roepnaam, geboortedatum, geslacht, team zaal/veld
- Opstellingen 2017-2018, 2019-2020: naam, geboortedatum, team, geslacht
- Opstellingen 2010-2016: alleen naam, team, geslacht (geen ID, geen geboortedatum)

Spelers zonder Sportlink Rel.nr krijgen code OW-0001, OW-0002, etc.
"""

import openpyxl
import xlrd
import json
import re
import os
import sys
from datetime import datetime
from collections import defaultdict

# Fix Windows console encoding voor speciale karakters (Turks, etc.)
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE = r"C:\Oranje Wit\docs\teamindelingen"
OUT = r"C:\Oranje Wit\data\spelers"
LEDENLIJST = r"C:\Oranje Wit\docs\Complete_Ledenlijst_TC.xlsx"

# --- Helpers ------------------------------------------------

STAFF_LABELS = {
    "trainers", "trainer", "coach", "begeleiders", "begeleider",
    "verzorgers", "verzorging", "team begeleiding", "techniek tr.",
    "hoofdtrainer", "assistent", "hoofd trainer", "scheidsrechter",
    "training / coaching", "staf", "eigen trainers",
}

SKIP_VALUES = {
    "naam", "dames", "heren", "opstelling", "", "vacature",
    "definitieve selectie na voorbereiding", "heer 1e selectie",
    "dame 1e selectie", "speelt 4 om 4", "speelt 5 om 5",
    "a1 speelt 4 om 4", "a2 speelt 5 om 5",
    "algemene reserves", "ar lijst",
}


def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, (int, float)):
        try:
            from xlrd import xldate_as_datetime
            return xldate_as_datetime(val, 0).strftime("%Y-%m-%d")
        except Exception:
            return None
    val = str(val).strip()
    if not val or val == "None":
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"]:
        try:
            return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("None", "0", "0.0", "#NAME?", "#REF!", "#VALUE!") else s


def normalize_name(naam):
    """Normaliseer naam voor matching: lowercase, geen tussenvoegsel-variatie."""
    if not naam:
        return ""
    naam = naam.strip().lower()
    naam = re.sub(r'\b[a-z]\.\s*', '', naam)
    naam = re.sub(r'\([^)]*\)', '', naam)
    for old, new in [("e\u0308", "e"), ("o\u0308", "o"), ("u\u0308", "u"),
                     ("\u00eb", "e"), ("\u00f6", "o"), ("\u00fc", "u"),
                     ("\u00e9", "e"), ("\u00e8", "e"), ("\u00ea", "e"),
                     ("\u00ef", "i"), ("\u00e1", "a"), ("\u00e4", "a"),
                     ("\u00f4", "o"), ("\u00e2", "a"), ("\u00fb", "u"),
                     ("\u00ee", "i"), ("\u00e7", "c"), ("\u00f1", "n"),
                     ("\u0327", ""), ("\u0302", ""), ("\u0301", ""),
                     ("\u0300", ""), ("\u0308", "")]:
        naam = naam.replace(old, new)
    naam = re.sub(r'\s+', ' ', naam).strip()
    naam = naam.replace(",", "").replace(".", "")
    return naam


def extract_achternaam_voornaam(naam_raw):
    """Parse diverse naamformaten naar (achternaam, voornaam/roepnaam)."""
    if not naam_raw:
        return ("", "")
    naam = naam_raw.strip()

    # Format: "Achternaam, V. (Voornaam)" of "Achternaam, V.Z.E. (Voornaam)"
    m = re.match(r'^(.+?),\s*(.+)$', naam)
    if m:
        achternaam = m.group(1).strip()
        rest = m.group(2).strip()
        rm = re.search(r'\(([^)]+)\)', rest)
        voornaam = rm.group(1) if rm else rest
        # Strip initials only when followed by a dot (e.g., "V." or "V.Z.E.")
        voornaam = re.sub(r'^([A-Z]\.)+\s*', '', voornaam).strip()
        if not voornaam:
            voornaam = rest
        return (achternaam, voornaam)

    # Format: "Achternaam (Voornaam)" without comma — e.g. "Valk (Jarno)", "Boeije (Thomas)"
    m2 = re.match(r'^(\S+)\s+\(([^)]+)\)$', naam)
    if m2:
        return (m2.group(1), m2.group(2))

    # Format: "Voornaam Achternaam" of "Voornaam van/de Achternaam"
    parts = naam.split()
    if len(parts) >= 2:
        tussenvoegsels = {"van", "de", "den", "der", "het", "ten", "ter", "in", "t", "'t", "vd", "v/d"}
        first = parts[0]
        rest_parts = parts[1:]
        achternaam_parts = []
        for i, p in enumerate(rest_parts):
            achternaam_parts.append(p)
            if p.lower() not in tussenvoegsels:
                achternaam_parts.extend(rest_parts[i+1:])
                break
        return (" ".join(achternaam_parts), first)

    return (naam, "")


def make_match_key(achternaam, voornaam):
    a = normalize_name(achternaam)
    v = normalize_name(voornaam)
    return f"{a}|{v}"


TUSSENVOEGSELS = {"van", "de", "den", "der", "het", "ten", "ter", "t", "'t", "in", "in't", "vd", "v/d"}


def _fuzzy_achternaam(ach):
    """Normaliseer achternaam voor fuzzy matching: ij/y, dubbele letters, etc."""
    s = normalize_name(ach)
    # Strip tussenvoegsels
    parts = [p for p in s.split() if p not in TUSSENVOEGSELS]
    s = " ".join(parts)
    # Nederlandse spellingvarianten
    s = s.replace("ij", "y")
    s = s.replace("ei", "ey")
    s = s.replace("oo", "o")
    s = s.replace("ee", "e")
    s = s.replace("aa", "a")
    s = s.replace("ck", "k")
    s = s.replace("ph", "f")
    s = s.replace("th", "t")
    s = s.replace("dt", "t")
    s = s.replace("sch", "s")
    # Dubbele medeklinkers reduceren (wall→wal, zwoll→zwol, koornneef→korneef)
    s = re.sub(r'(.)\1', r'\1', s)
    # Achterliggende s verwijderen (dekkers→dekker)
    if s.endswith("s") and len(s) > 3:
        s = s[:-1]
    return s


def make_fuzzy_key(achternaam, voornaam):
    """Maak een fuzzy match key op basis van genormaliseerde achternaam + voornaam."""
    a = _fuzzy_achternaam(achternaam)
    v_norm = normalize_name(voornaam) if voornaam else ""
    v = v_norm.split()[0] if v_norm.split() else ""
    if not a or not v:
        return ""
    return f"~{a}|{v}"


def normalize_team(team_str):
    if not team_str:
        return None
    t = team_str.strip()
    if not t or t.lower() in ("-", "stopt", "ar", "gestopt", "?", "x1", ""):
        return None
    # Remove patterns like "S6x1", "A3x1" -> "S6", "A3"
    m = re.match(r'^([A-Z]\d+)x\d+$', t, re.IGNORECASE)
    if m:
        t = m.group(1)
    t = re.sub(r'^OW\s+', '', t)
    if t.lower() in ("sen", "senioren", "s"):
        return "Senioren"
    return t


def categorize_team(team):
    if not team:
        return None
    t = team.upper()
    if t == "SENIOREN" or t.startswith("MW") or t.startswith("JOW"):
        return "Senioren"
    if re.match(r'^S\d', t):
        return "Senioren"
    if t.startswith("A"):
        return "A-Junioren"
    if t.startswith("B"):
        return "B-Aspiranten"
    if t.startswith("C"):
        return "C-Aspiranten"
    if t.startswith("D"):
        return "D-Pupillen"
    if t.startswith("E"):
        return "E-Pupillen"
    if t.startswith("F"):
        return "F-Pupillen"
    if t.startswith("K"):
        return "Kangoeroes"
    if t.startswith("U1"):
        return "A-Junioren"
    if t.startswith("J") or t.startswith("OW J"):
        return "Jeugd"
    return "Senioren"


def _detect_categorie(sheet_name):
    """Bepaal leeftijdscategorie op basis van sheet-naam."""
    s = sheet_name.lower()
    if "senior" in s or s in ("s", "team sen"):
        return "Senioren"
    if "junior" in s or s in ("a", "team a"):
        return "A-Junioren"
    if "aspirant" in s and "b" in s or s in ("b", "team b"):
        return "B-Aspiranten"
    if "aspirant" in s and "c" in s or s in ("c", "team c"):
        return "C-Aspiranten"
    if "pupil" in s and "d" in s or s in ("d", "team d"):
        return "D-Pupillen"
    if "pupil" in s and "e" in s or s in ("e", "team e"):
        return "E-Pupillen"
    if "pupil" in s and "f" in s or s in ("f", "f+k", "team f"):
        return "F-Pupillen"
    if "kangoeroe" in s:
        return "Kangoeroes"
    if "recreant" in s or s == "ar en recreanten":
        return "Senioren"
    return sheet_name


def is_name(text):
    """Heuristiek: is dit een persoonsnaam?"""
    if not text or len(text) < 3:
        return False
    t_low = text.lower().strip()
    if t_low in SKIP_VALUES or t_low in STAFF_LABELS:
        return False
    if re.match(r'^\d+\.?\d*$', text):
        return False
    # Staff/label keywords anywhere in the text
    if any(kw in t_low for kw in [
        "trainer", "coach", "begeleider", "scheidsrechter", "verzorg",
        "opstelling", "selectie", "klasse", "hoofdklasse", "hoofd klasse",
        "vacature", "coordinat", "coördinat", "techniek",
        "manager", "oranje wit", "carrousel", "tc-lid",
    ]):
        return False
    # Team-label patterns: "S5 S6", "B1   4/4", "B2  5/4", "Veld Team 6/7/8/9", "Heren ..."
    if re.match(r'^[A-Z]\d\s+[A-Z\d/]', text):
        return False
    if re.match(r'^(?:Veld|Zaal)\s+Team\b', text, re.IGNORECASE):
        return False
    if re.match(r'^(?:Heren|Dames|Heer|Dame)\s+', text):
        return False
    # "Team X" labels: "Team B1", "Team 1 en 2", "Team A1 en A2"
    if re.match(r'^Team\s+', text, re.IGNORECASE):
        return False
    # Competition formats: "A1/A2   4/4 + 5/5", "S3/S4  5/5 + 5/5"
    if re.search(r'\d/\d\s*\+?\s*\d/\d', text):
        return False
    # Must start with a capital letter
    if not re.match(r'^[A-Z]', text):
        return False
    # Must have comma or multiple words
    if "," in text or len(text.split()) >= 2:
        return True
    return False


# --- Speler registry ----------------------------------------

class SpelerRegistry:
    def __init__(self):
        self.by_rel_nr = {}
        self.by_match_key = {}
        self.by_fuzzy_key = {}
        self.ow_counter = 0

    def _new_speler(self, speler_id, naam, roepnaam, geslacht, geboortedatum, bron):
        return {
            "speler_id": speler_id,
            "naam": naam,
            "roepnaam": roepnaam,
            "geslacht": geslacht,
            "geboortedatum": geboortedatum,
            "bron_id": bron,
            "seizoenen": {},
            "gestopt": None
        }

    def _next_ow_id(self):
        self.ow_counter += 1
        return f"OW-{self.ow_counter:04d}"

    def _register_keys(self, speler_id, naam, roepnaam):
        ach, vn = extract_achternaam_voornaam(naam)
        key = make_match_key(ach, vn)
        if key and key != "|":
            self.by_match_key[key] = speler_id
        if roepnaam:
            key2 = make_match_key(ach, roepnaam)
            if key2 and key2 != "|":
                self.by_match_key[key2] = speler_id
        # Register simplified key: achternaam + first word of voornaam only
        # This helps match "Jong, de B (Bas)" with "Jong, Bas de"
        for v in [vn, roepnaam]:
            if v:
                first_word = v.split()[0] if v.split() else ""
                if first_word and first_word.lower() not in TUSSENVOEGSELS:
                    key3 = make_match_key(ach, first_word)
                    if key3 and key3 != "|":
                        self.by_match_key[key3] = speler_id
        # Register key with achternaam-root (strip leading tussenvoegsels)
        # "van der Wall" -> also register "wall|jeffrey"
        ach_parts = ach.split() if ach else []
        if len(ach_parts) > 1:
            root = " ".join(p for p in ach_parts if p.lower() not in TUSSENVOEGSELS)
            if root and root != ach:
                for v in [vn, roepnaam]:
                    if v:
                        first_word = v.split()[0] if v.split() else v
                        if first_word and first_word.lower() not in TUSSENVOEGSELS:
                            rkey = make_match_key(root, first_word)
                            if rkey and rkey != "|":
                                self.by_match_key[rkey] = speler_id
        # Register key with first part of hyphenated achternaam
        # "Kleingeld-Wibbens" -> also register "kleingeld|liesbeth"
        if ach and "-" in ach:
            first_part = ach.split("-")[0].strip()
            if first_part:
                for v in [vn, roepnaam]:
                    if v:
                        first_word = v.split()[0] if v.split() else v
                        if first_word and first_word.lower() not in TUSSENVOEGSELS:
                            hkey = make_match_key(first_part, first_word)
                            if hkey and hkey != "|":
                                self.by_match_key[hkey] = speler_id
        # Register fuzzy achternaam keys (vangt ij/y, dubbele letters, etc.)
        for v in [vn, roepnaam]:
            if v:
                fkey = make_fuzzy_key(ach, v)
                if fkey:
                    self.by_fuzzy_key[fkey] = speler_id

    def add_with_rel_nr(self, rel_nr, naam, roepnaam, geslacht, geboortedatum):
        if rel_nr not in self.by_rel_nr:
            self.by_rel_nr[rel_nr] = self._new_speler(
                rel_nr, naam, roepnaam, geslacht, geboortedatum, "sportlink"
            )
        else:
            # Always update to latest non-empty values (Supersheets processed chronologically)
            sp = self.by_rel_nr[rel_nr]
            if geboortedatum:
                sp["geboortedatum"] = geboortedatum
            if roepnaam:
                sp["roepnaam"] = roepnaam
            if geslacht:
                sp["geslacht"] = geslacht
            if naam:
                sp["naam"] = naam
        self._register_keys(rel_nr, naam, roepnaam)
        return rel_nr

    def find_or_create(self, naam, roepnaam, geslacht, geboortedatum):
        ach, vn = extract_achternaam_voornaam(naam)

        # Build list of achternaam variants to try
        ach_variants = [ach]
        ach_parts = ach.split() if ach else []
        if len(ach_parts) > 1:
            root = " ".join(p for p in ach_parts if p.lower() not in TUSSENVOEGSELS)
            if root and root != ach:
                ach_variants.append(root)

        def _try_match(key):
            if key in self.by_match_key:
                rel = self.by_match_key[key]
                sp = self.by_rel_nr[rel]
                if geboortedatum and not sp["geboortedatum"]:
                    sp["geboortedatum"] = geboortedatum
                if geslacht and not sp["geslacht"]:
                    sp["geslacht"] = geslacht
                return rel
            return None

        # Try exact match key (including achternaam-root variant)
        for a in ach_variants:
            for try_vn in [vn, roepnaam]:
                if try_vn:
                    rel = _try_match(make_match_key(a, try_vn))
                    if rel:
                        return rel

        # Try simplified key: achternaam + first word of voornaam
        for a in ach_variants:
            for try_vn in [vn, roepnaam]:
                if try_vn:
                    first_word = try_vn.split()[0] if try_vn.split() else ""
                    if first_word and first_word.lower() not in TUSSENVOEGSELS:
                        rel = _try_match(make_match_key(a, first_word))
                        if rel:
                            return rel

        # Try fuzzy achternaam match (vangt ij/y, dubbele letters, trailing s, etc.)
        for try_vn in [vn, roepnaam]:
            if try_vn:
                fkey = make_fuzzy_key(ach, try_vn)
                if fkey and fkey in self.by_fuzzy_key:
                    rel = self.by_fuzzy_key[fkey]
                    sp = self.by_rel_nr[rel]
                    if geboortedatum and not sp["geboortedatum"]:
                        sp["geboortedatum"] = geboortedatum
                    if geslacht and not sp["geslacht"]:
                        sp["geslacht"] = geslacht
                    return rel

        # Try geboortedatum + achternaam (exacte achternaam)
        if geboortedatum:
            norm_ach = normalize_name(ach)
            for rel, sp in self.by_rel_nr.items():
                if sp["geboortedatum"] == geboortedatum:
                    sp_ach, _ = extract_achternaam_voornaam(sp["naam"])
                    if normalize_name(sp_ach) == norm_ach:
                        key = make_match_key(ach, vn)
                        self.by_match_key[key] = rel
                        return rel

        # Try geboortedatum + fuzzy achternaam
        if geboortedatum:
            fuzzy_ach = _fuzzy_achternaam(ach)
            for rel, sp in self.by_rel_nr.items():
                if sp["geboortedatum"] == geboortedatum:
                    sp_ach, _ = extract_achternaam_voornaam(sp["naam"])
                    if _fuzzy_achternaam(sp_ach) == fuzzy_ach:
                        key = make_match_key(ach, vn)
                        self.by_match_key[key] = rel
                        return rel

        # No match - create OW code
        ow_id = self._next_ow_id()
        display_naam = naam if naam else f"{ach}, {vn}"
        self.by_rel_nr[ow_id] = self._new_speler(
            ow_id, display_naam, roepnaam or vn, geslacht, geboortedatum, "name-match"
        )
        self._register_keys(ow_id, display_naam, roepnaam or vn)
        return ow_id

    def add_season(self, speler_id, seizoen, team, categorie=None, rol="speler"):
        if speler_id not in self.by_rel_nr:
            return
        team_norm = normalize_team(team)
        if not team_norm:
            return
        cat = categorie or categorize_team(team_norm)
        sp = self.by_rel_nr[speler_id]
        if seizoen not in sp["seizoenen"]:
            sp["seizoenen"][seizoen] = {"team": team_norm, "categorie": cat, "rol": rol}

    def mark_stopped(self, speler_id, seizoen):
        if speler_id in self.by_rel_nr:
            self.by_rel_nr[speler_id]["gestopt"] = seizoen


# --- Parsers ------------------------------------------------

def _parse_teams_sheet(wb, seizoen, registry):
    """Parse de 'Teams' sheet en update teamindelingen in de registry.

    Gebruikt voor 2021-2022 waar de Spelerslijst geen kolom met het
    huidige-seizoen team bevat. De Teams sheet heeft secties per team
    met dames (col 1-3) en heren (col 4-6) naast elkaar.
    """
    if "Teams" not in wb.sheetnames:
        print("    Teams sheet: niet gevonden")
        return

    ws = wb["Teams"]
    rows = list(ws.iter_rows(values_only=True))

    current_team = None
    updates = 0
    not_matched = []

    staff_kw = {"coach", "trainer", "assistent", "fysio", "verzorger",
                "begeleider", "verzorging", "scheidsrechter"}

    for row in rows:
        cells = [str(c).strip() if c else ""
                 for c in (row[:8] if len(row) >= 8
                           else list(row) + [""] * (8 - len(row)))]

        # Team-header: col[0] bevat categorie, col[1] bevat teamnaam
        if cells[0] and cells[1]:
            current_team = cells[1]
            continue

        # Gender-count rij ("Dames (N)" / "Heren (N)") → skip
        if re.match(r'^(Dames|Heren)\s*\(', cells[1]):
            continue

        # Lege rij → skip
        if not any(cells[1:6]):
            continue

        # Staff rij → skip
        if any(kw in cells[1].lower() for kw in staff_kw):
            continue

        if not current_team:
            continue

        # Speler-rijen: dames col[1]+[2], heren col[4]+[5]
        for name_col, roep_col in [(1, 2), (4, 5)]:
            naam_raw = cells[name_col]
            roepnaam = cells[roep_col]
            if not naam_raw or not roepnaam:
                continue

            # Achternaam extraheren uit "Achternaam, Init." formaat
            ach = naam_raw.split(",")[0].strip() if "," in naam_raw else naam_raw.strip()

            # Zoek speler in registry via match_key (achternaam + roepnaam)
            key = make_match_key(ach, roepnaam)
            speler_id = registry.by_match_key.get(key)

            if not speler_id:
                fkey = make_fuzzy_key(ach, roepnaam)
                speler_id = registry.by_fuzzy_key.get(fkey) if fkey else None

            if speler_id and speler_id in registry.by_rel_nr:
                sp = registry.by_rel_nr[speler_id]
                new_team = normalize_team(current_team)
                if not new_team:
                    continue
                new_cat = categorize_team(new_team)
                if seizoen in sp["seizoenen"]:
                    sp["seizoenen"][seizoen]["team"] = new_team
                    sp["seizoenen"][seizoen]["categorie"] = new_cat
                else:
                    sp["seizoenen"][seizoen] = {
                        "team": new_team, "categorie": new_cat, "rol": "speler"
                    }
                updates += 1
            else:
                not_matched.append(f"{naam_raw} ({roepnaam}) → {current_team}")

    print(f"    Teams sheet: {updates} team-updates, {len(not_matched)} niet-gematcht")
    if not_matched:
        for nm in not_matched[:10]:
            print(f"      ? {nm}")
        if len(not_matched) > 10:
            print(f"      ... en {len(not_matched) - 10} meer")


def parse_supersheets(registry):
    """Parse Supersheets 2021-2026."""
    files = [
        ("2021-2022", os.path.join(BASE, "Supersheet indeling 1.0 - 2021-2022.xlsx")),
        ("2022-2023", os.path.join(BASE, "Supersheet indeling 1.0 - 2022-2023.xlsx")),
        ("2023-2024", os.path.join(BASE, "Supersheet indeling 1.0 - 2023-2024.xlsx")),
        ("2024-2025", os.path.join(BASE, "Supersheet indeling 1.0 - 2024-2025.xlsx")),
        ("2025-2026", os.path.join(BASE, "Supersheet indeling 1.0 - 2025-2026.xlsx")),
    ]

    for seizoen, path in files:
        print(f"\n  Supersheet {seizoen}...")
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

        if "Spelerslijst" in wb.sheetnames:
            ws = wb["Spelerslijst"]
            rows = list(ws.iter_rows(values_only=True))
            header = [clean(c) for c in rows[0][:25]]
            rel_cols = [i for i, h in enumerate(header) if "rel" in h.lower() and "nr" in h.lower()]

            count = 0
            for row in rows[1:]:
                cells = list(row[:25]) if len(row) >= 25 else list(row) + [None] * (25 - len(row))
                for side_idx, rel_col in enumerate(rel_cols):
                    rel_nr = clean(cells[rel_col]) if cells[rel_col] else ""
                    if not rel_nr or not rel_nr.startswith("N"):
                        continue
                    naam = clean(cells[rel_col + 1]) if cells[rel_col + 1] else ""
                    roepnaam = clean(cells[rel_col + 2]) if cells[rel_col + 2] else ""
                    gebdat = parse_date(cells[rel_col + 4]) if len(cells) > rel_col + 4 else None
                    geslacht = "V" if side_idx == 0 else "M"
                    team = ""
                    # Detecteer juiste teamkolom per side:
                    # "Team op Peildatum" (2021-2025) = Sportlink-categorieletter → skip, echte indeling op +6
                    # "Team [seizoen]" (2025-2026+) = directe teamindeling → gebruik +5
                    header_at_5 = header[rel_col + 5] if len(header) > rel_col + 5 else ""
                    if "peildatum" in header_at_5.lower():
                        team_col = rel_col + 6
                    else:
                        team_col = rel_col + 5
                    if len(cells) > team_col and cells[team_col]:
                        team = clean(cells[team_col])
                    if not naam:
                        continue
                    sid = registry.add_with_rel_nr(rel_nr, naam, roepnaam, geslacht, gebdat)
                    registry.add_season(sid, seizoen, team)
                    count += 1
            print(f"    Spelerslijst: {count} speler-seizoenen")

        # 2021-2022: Spelerslijst heeft geen huidige-seizoen teamkolom,
        # alleen "Team 20-21" (vorig jaar). Overschrijf met Teams sheet.
        if seizoen == "2021-2022":
            _parse_teams_sheet(wb, seizoen, registry)

        if "Stoppers" in wb.sheetnames:
            ws = wb["Stoppers"]
            rows = list(ws.iter_rows(values_only=True))
            stop_count = 0
            for row in rows[1:]:
                cells = [clean(c) for c in (row[:20] if len(row) >= 20 else list(row) + [None] * 20)]
                for c in cells:
                    if c.startswith("N") and len(c) >= 6 and re.match(r'^N[A-Z0-9]{5,}$', c):
                        registry.mark_stopped(c, seizoen)
                        stop_count += 1
            print(f"    Stoppers: {stop_count}")

        wb.close()


A2_FILES = [
    ("2018-2019", "A2 zwt - Oranje Wit (D) 2018 2019.xlsm"),
    ("2019-2020", "A2 zwt - Oranje Wit (D) 06-2020.xlsm"),
    ("2020-2021", "A2 mix - Oranje Wit (D)  09-2021.xlsm"),
    ("2021-2022", "A2 mix - Oranje Wit (D) 20-2022.xlsm"),
]


def parse_a2_formulier(registry, seizoen, filename):
    """Parse één A2 KNKV-formulier (.xlsm).

    Kolomstructuur Leden-sheet (rij 4 = header, data vanaf rij 5):
    2=zaal team, 3=veld team, 5=geslacht, 6=rel_nr, 7=naam, 8=roepnaam, 9=gebdat
    """
    path = os.path.join(BASE, filename)
    print(f"\n  A2 formulier {seizoen} ({filename})...")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True, keep_vba=False)
    ws = wb["Leden"]
    rows = list(ws.iter_rows(values_only=True))

    count = 0
    skipped_no_rel = 0
    for row in rows[4:]:
        cells = list(row[:20]) if len(row) >= 20 else list(row) + [None] * (20 - len(row))
        rel_nr = clean(cells[6])
        if not rel_nr or not rel_nr.startswith("N"):
            naam_check = clean(cells[7])
            if naam_check:
                skipped_no_rel += 1
            continue
        naam = clean(cells[7])
        roepnaam = clean(cells[8])
        gebdat = parse_date(cells[9])
        geslacht_raw = clean(cells[5])
        geslacht = "V" if geslacht_raw == "V" else "M" if geslacht_raw == "M" else ""
        team_zaal = clean(cells[2])
        team_veld = clean(cells[3])
        team = team_zaal or team_veld
        if not naam:
            continue
        sid = registry.add_with_rel_nr(rel_nr, naam, roepnaam, geslacht, gebdat)
        registry.add_season(sid, seizoen, team)
        count += 1

    wb.close()
    print(f"    Leden: {count} spelers (met Rel.nr)")
    if skipped_no_rel:
        print(f"    Overgeslagen: {skipped_no_rel} zonder Rel.nr")


def parse_all_a2(registry):
    """Parse alle A2 formulieren (2018-2022)."""
    for seizoen, filename in A2_FILES:
        parse_a2_formulier(registry, seizoen, filename)


def _parse_numbered_rows(cells, raw_cells, current_gender, has_gebdat, registry, seizoen, current_team, categorie, heren_col=None, dames_col=None):
    """Parse genummerde spelersrijen: '1', 'Naam', eventueel 'gebdat'.
    Returns number of players found."""
    count = 0

    # Zoek paren van (nummer, naam) in de cellen
    for i in range(len(cells) - 1):
        c = cells[i]
        if not re.match(r'^\d+\.?\d*$', c):
            continue
        naam = cells[i + 1]
        if not naam or not is_name(naam):
            continue

        # Determine gender from column position (side-by-side layout) or fallback
        if heren_col is not None and dames_col is not None:
            gender = "M" if abs(i - heren_col) < abs(i - dames_col) else "V"
        else:
            gender = current_gender

        gebdat = None
        if has_gebdat and i + 2 < len(raw_cells):
            gebdat = parse_date(raw_cells[i + 2])

        # Check for team code after name
        team_code = cells[i + 2] if i + 2 < len(cells) and not has_gebdat else ""
        if team_code and re.match(r'^[A-Z]\d', team_code):
            team = team_code
        else:
            team = current_team or (categorie[0] if categorie else "")

        ach, vn = extract_achternaam_voornaam(naam)
        if ach and len(ach) > 1:
            sid = registry.find_or_create(naam, vn, gender, gebdat)
            registry.add_season(sid, seizoen, team, categorie)
            count += 1

    return count


def parse_xlsx_opstelling(registry, seizoen, path, include_sheets=None, skip_sheets=None):
    """Parse .xlsx opstellingen (2012-2013, 2016-2017, 2017-2018, 2019-2020, 2020-2021).

    include_sheets: als gezet, alleen deze sheets parsen
    skip_sheets: als gezet, deze sheets overslaan
    """
    print(f"\n  Opstelling {seizoen}...")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    count = 0
    for sname in wb.sheetnames:
        if include_sheets and sname not in include_sheets:
            continue
        if skip_sheets and sname in skip_sheets:
            continue
        ws = wb[sname]
        rows = list(ws.iter_rows(values_only=True))
        categorie = _detect_categorie(sname)

        # Detect if this sheet has geboortedata
        has_gebdat = False
        for r in rows[:6]:
            for c in (r[:15] if len(r) >= 15 else list(r)):
                if c and "Geb" in str(c) and "Dat" in str(c):
                    has_gebdat = True
                    break

        current_team = ""
        current_gender = ""
        heren_col = None
        dames_col = None
        in_staff_section = False
        in_player_section = False

        for row in rows:
            cells = [clean(c) for c in (row[:15] if len(row) >= 15 else list(row) + [None] * 15)]
            raw_cells = list(row[:15]) if len(row) >= 15 else list(row) + [None] * 15

            # Detect team header
            for c in cells:
                m = re.match(
                    r'^(?:Team\s+)?(S\d|A\d|B\d|C\d|D\d|E\d|F\d|K\d|MW\d|JOW|'
                    r'S1S2|A1A2|S1/S2|A1/A2|B1/B2|C1/C2|D1/D2|'
                    r'\d+\s+en\s+\d+)\b', c, re.IGNORECASE
                )
                if m:
                    raw = m.group(1).replace("/", "")
                    if "en" in raw:
                        current_team = categorie[0] + "1" if categorie else raw
                    else:
                        current_team = raw
                    in_staff_section = False
                    in_player_section = False
                    heren_col = None
                    dames_col = None
                    break

            # Detect staff label (startswith for "Trainers A1", "Trainer/coach" etc.)
            for c in cells:
                c_low = c.lower()
                if c_low in STAFF_LABELS or any(c_low.startswith(sl) for sl in STAFF_LABELS):
                    in_staff_section = True
                    break

            # Detect gender header (track column positions for side-by-side layout)
            row_heren = None
            row_dames = None
            for ci, c in enumerate(cells):
                if c.lower() == "heren":
                    row_heren = ci
                elif c.lower() == "dames":
                    row_dames = ci
            if row_heren is not None or row_dames is not None:
                in_player_section = True
                in_staff_section = False
                if row_heren is not None:
                    heren_col = row_heren
                    current_gender = "M"
                if row_dames is not None:
                    dames_col = row_dames
                    current_gender = "V"
                # If both on same row, current_gender not meaningful — use column positions
                if row_heren is not None and row_dames is not None:
                    current_gender = ""

            # Skip rows before Heren/Dames header (trainers, staff, etc.)
            if not in_player_section:
                continue

            if has_gebdat:
                # Format: '', nr, 'Naam', 'YYYY-MM-DD', leeftijd, nr, 'Naam', 'YYYY-MM-DD', leeftijd
                # Determine gender per side from column positions
                if heren_col is not None and dames_col is not None:
                    sides = [(1, "M" if heren_col < dames_col else "V"),
                             (5, "V" if heren_col < dames_col else "M")]
                else:
                    sides = [(1, "V"), (5, "M")]
                for side_start, gender in sides:
                    if side_start + 2 >= len(cells):
                        continue
                    nr = cells[side_start]
                    naam = cells[side_start + 1]
                    if not naam or not nr.replace(".", "").isdigit():
                        continue
                    if not is_name(naam):
                        continue
                    gebdat = parse_date(raw_cells[side_start + 2])
                    ach, vn = extract_achternaam_voornaam(naam)
                    sid = registry.find_or_create(naam, vn, gender, gebdat)
                    team = current_team or (categorie[0] if categorie else "")
                    registry.add_season(sid, seizoen, team, categorie)
                    count += 1
            else:
                # Try numbered rows (pass column positions for side-by-side gender)
                n = _parse_numbered_rows(cells, raw_cells, current_gender, False,
                                        registry, seizoen, current_team, categorie,
                                        heren_col=heren_col, dames_col=dames_col)
                count += n

                # If no numbered rows, try loose name detection
                # Limit to player columns (exclude trainer columns beyond Heren/Dames area)
                if n == 0:
                    max_col = max(heren_col, dames_col) + 2 if heren_col is not None and dames_col is not None else 6
                    for ci, c in enumerate(cells[:max_col]):
                        if is_name(c):
                            ach, vn = extract_achternaam_voornaam(c)
                            if ach and len(ach) > 1:
                                if heren_col is not None and dames_col is not None:
                                    gender = "M" if abs(ci - heren_col) < abs(ci - dames_col) else "V"
                                else:
                                    gender = current_gender
                                sid = registry.find_or_create(c, vn, gender, None)
                                team = current_team or (categorie[0] if categorie else "")
                                registry.add_season(sid, seizoen, team, categorie)
                                count += 1

    wb.close()
    print(f"    Totaal: {count} speler-seizoenen")


def parse_xls_opstelling(registry, seizoen, path):
    """Parse .xls opstellingen 2010-2016.

    Structure: sheets per leeftijdscategorie. Trainers/coaches listed BEFORE
    Heren/Dames headers. Players in numbered rows AFTER Heren/Dames headers.
    """
    print(f"\n  Opstelling {seizoen}...")
    wb = xlrd.open_workbook(path)

    count = 0
    for sname in wb.sheet_names():
        ws = wb.sheet_by_name(sname)
        categorie = _detect_categorie(sname)

        # Pre-scan: detect if this sheet uses Heren/Dames headers
        sheet_has_gender_headers = False
        for r in range(ws.nrows):
            for c_idx in range(min(ws.ncols, 12)):
                val = str(ws.cell_value(r, c_idx)).strip().lower()
                if val in ("heren", "dames"):
                    sheet_has_gender_headers = True
                    break
            if sheet_has_gender_headers:
                break

        current_team = ""
        current_gender = ""
        heren_col = None
        dames_col = None
        in_staff_section = False
        in_player_section = False

        for r in range(ws.nrows):
            cells = [clean(str(ws.cell_value(r, c_idx)))
                     for c_idx in range(min(ws.ncols, 12))]

            # Detect team header
            for c in cells:
                m = re.match(
                    r'^(?:Team\s+)?(S\d|A\d|B\d|C\d|D\d|E\d|F\d|K\d|MW\d|'
                    r'S1S2|A1A2|S1/S2|A1/A2|B1/B2|C1/C2|D1/D2|'
                    r'\d+\s+en\s+\d+)\b', c, re.IGNORECASE
                )
                if m:
                    raw = m.group(1).replace("/", "")
                    if "en" in raw:
                        current_team = categorie[0] + "1" if categorie else raw
                    else:
                        current_team = raw
                    in_staff_section = False
                    in_player_section = False
                    heren_col = None
                    dames_col = None
                    break

            if sheet_has_gender_headers:
                # 2010-2012 layout: Trainers before Heren/Dames, players after
                for c in cells:
                    c_low = c.lower()
                    if c_low in STAFF_LABELS or any(c_low.startswith(sl) for sl in STAFF_LABELS):
                        in_staff_section = True
                        break

                # Detect gender header
                row_heren = None
                row_dames = None
                for ci, c in enumerate(cells):
                    if c.lower() == "heren":
                        row_heren = ci
                    elif c.lower() == "dames":
                        row_dames = ci
                if row_heren is not None or row_dames is not None:
                    in_player_section = True
                    in_staff_section = False
                    if row_heren is not None:
                        heren_col = row_heren
                        current_gender = "M"
                    if row_dames is not None:
                        dames_col = row_dames
                        current_gender = "V"
                    if row_heren is not None and row_dames is not None:
                        current_gender = ""

                # Skip rows before Heren/Dames header
                if not in_player_section:
                    continue
            # else: 2013-2016 layout without Heren/Dames — parse columns 0-5 directly

            # Parse numbered player rows
            n = _parse_numbered_rows(cells, cells, current_gender, False,
                                    registry, seizoen, current_team, categorie,
                                    heren_col=heren_col, dames_col=dames_col)
            count += n

            # If no numbered rows, try loose name detection (only player columns 0-5)
            if n == 0:
                for ci, c in enumerate(cells[:6]):
                    if is_name(c):
                        ach, vn = extract_achternaam_voornaam(c)
                        if ach and len(ach) > 1:
                            if heren_col is not None and dames_col is not None:
                                gender = "M" if abs(ci - heren_col) < abs(ci - dames_col) else "V"
                            else:
                                gender = current_gender
                            sid = registry.find_or_create(c, vn, gender, None)
                            team = current_team or (categorie[0] if categorie else "")
                            registry.add_season(sid, seizoen, team, categorie)
                            count += 1

    print(f"    Totaal: {count} speler-seizoenen")


# --- Deduplicatie -------------------------------------------

# Handmatige merge-tabel: (primary, secondary)
# primary: Sportlink ID, of "naam:Achternaam, Voornaam" voor naam-lookup
# secondary: Sportlink ID, of ("achternaam", "voornaam") tuple voor naam-lookup
#
# Naam-gebaseerde lookup is robuust tegen verschuivende OW-IDs.
# Entries die fuzzy matching nu automatisch afhandelt zijn VERWIJDERD:
#   Dekkers/Dekker, Wall/Wal, Zwoll/Zwol, Ooyen/Ooijen (trailing s / dubbele letters / ij-y)
#
MERGE_TABLE = [
    # --- Sportlink <> Sportlink (stabiele IDs) ---
    ("NJC17J5", "NJC17J6"),      # Gent, Lex van <- Rel.nr off by 1
    ("NJF85B1", "NKW81C0"),      # Klink, Xanne <- Klink, X (initials)
    ("NKS24S7", "NMJ81C1"),      # Wijngaarden, Bente van <- initials
    ("NKK27G4", "NML43T8"),      # Valk, Marlies <- Valk, M.E. (initials)
    ("NLJ04Z4", "NLJ4Z4"),       # Lohuis, Floor <- Rel.nr missing 0
    ("NLR34F7", "NMF58C3"),      # Jongkind, Floortje <- initials
    ("NLX64V0", "NMK07N7"),      # Boer, Esmee de <- initials
    ("NMJ53F3", "NMJ63F3"),      # Eva van Rooij <- Rel.nr off by 1
    ("NMG13P0", "NLD55M5"),      # Boer, Layla de <- initials
    ("NMH37W2", "NMJ14Y9"),      # Linden, Ize van der <- initials
    ("NMJ03L0", "NMJ3L0"),       # Levi Everaarts <- Rel.nr missing 0
    ("NMZ29T0", "NMZ29TO"),      # Branley van Wingerden <- O vs 0 typo

    # --- Sportlink <- OW naam-match (getrouwd / andere achternaam / niet-fuzzy) ---
    ("NJC15T9", ("Visser", "Diana")),              # Visser-Michelsen <- Diana Visser
    ("NJC15T9", ("Michelsen", "Diana")),            # Visser-Michelsen <- Diana Michelsen
    ("NHT58X9", ("Tol", "Moniek")),                 # Talens, Moniek <- Tol (getrouwd)
    ("NKF94H2", ("Fernout", "Tirsa")),              # Fernhout <- Fernout (h verschil)
    ("NGD97S8", ("in't Veld", "Lisette")),          # Veld, in 't <- spatie-variant
    ("NFW28D3", ("in't Veld", "Dorien")),           # Veld, in 't <- spatie-variant
    ("NLX08G7", ("Kamezman", "Zoe")),               # Kamerman, Zoe <- Kamezman (typo)
    ("NKD20B1", ("Zuiderwijk", "Ilse")),            # Zuijderwijk <- spelling
    ("NKK88N6", ("St. Nicolaas", "Marcel")),        # Sint Nicolaas <- afkorting
    ("NFX24T4", ("Haksteen", "Tamara")),            # Opstal-Haksteen <- meisjesnaam
    ("NJC15N7", ("Leijs", "Renee")),                # Van Haren-Leijs <- trouwnaam
    ("NFW40Y6", ("Dijk", "Mariette")),              # Grootenboer-Dijk <- trouwnaam
    ("NGL91R1", ("Haren", "Simone")),               # Kuijper-Vanhooydonck <- van Haren
    ("NJB68C3", ("Bruin", "Daphne")),               # Bruyn, de <- Bruin (ui/uy)
    ("NLL20B6", ("Wenzen", "Maud")),                # Wensen, van <- Wenzen (z/s)
    ("NLG58Y1", ("Rubio", "Estelle")),              # Rubio Cubillos <- deelachternaam

    # --- OW <- OW naam-match (twee OW-entries voor dezelfde persoon) ---
    ("naam:Meuhlhaus, Babette", ("Meuhlhaus", "Babet")),
    ("naam:Kardinaal, Leo", ("Kardienaal", "Leo")),
]


def _resolve_id(registry, spec):
    """Zoek een speler op ID of op (achternaam, voornaam) tuple.
    Returns speler_id of None."""
    if isinstance(spec, str):
        # "naam:Achternaam, Voornaam" — naam-gebaseerde lookup
        if spec.startswith("naam:"):
            naam = spec[5:]
            ach, vn = extract_achternaam_voornaam(naam)
        else:
            # Direct ID
            return spec if spec in registry.by_rel_nr else None
    elif isinstance(spec, tuple):
        ach, vn = spec
    else:
        return None

    # Zoek op exacte match key
    key = make_match_key(ach, vn)
    if key in registry.by_match_key:
        sid = registry.by_match_key[key]
        if sid in registry.by_rel_nr:
            return sid
    # Probeer simplified key (eerste woord voornaam)
    first_vn = vn.split()[0] if vn and vn.split() else ""
    if first_vn:
        skey = make_match_key(ach, first_vn)
        if skey in registry.by_match_key:
            sid = registry.by_match_key[skey]
            if sid in registry.by_rel_nr:
                return sid
    # Probeer fuzzy key
    fkey = make_fuzzy_key(ach, vn)
    if fkey and fkey in registry.by_fuzzy_key:
        sid = registry.by_fuzzy_key[fkey]
        if sid in registry.by_rel_nr:
            return sid
    return None


def merge_duplicates(registry):
    """Voeg bekende duplicaten samen op basis van de merge-tabel."""
    merged = 0
    for primary_spec, secondary_spec in MERGE_TABLE:
        primary_id = _resolve_id(registry, primary_spec)
        if not primary_id:
            continue
        secondary_id = _resolve_id(registry, secondary_spec)
        if not secondary_id or secondary_id == primary_id:
            continue

        primary = registry.by_rel_nr[primary_id]
        secondary = registry.by_rel_nr[secondary_id]
        # Merge seizoenen (primary wint bij overlap)
        for sz, info in secondary["seizoenen"].items():
            if sz not in primary["seizoenen"]:
                primary["seizoenen"][sz] = info
        # Vul ontbrekende velden aan
        if not primary["geboortedatum"] and secondary["geboortedatum"]:
            primary["geboortedatum"] = secondary["geboortedatum"]
        if not primary["geslacht"] and secondary["geslacht"]:
            primary["geslacht"] = secondary["geslacht"]
        # Verwijder secondary
        del registry.by_rel_nr[secondary_id]
        merged += 1
    return merged


# --- Ledenlijst verrijking ----------------------------------

def enrich_from_ledenlijst(registry):
    """Verrijk OW-spelers met Sportlink ID uit Complete Ledenlijst.

    1. Match OW-spelers op geboortedatum + achternaam
    2. Match OW-spelers op unieke naam-key
    3. Upgrade matched OW-spelers naar Sportlink Rel.nr
    4. Verrijk bestaande spelers met ontbrekend geslacht/geboortedatum
    """
    if not os.path.exists(LEDENLIJST):
        print("  Ledenlijst niet gevonden, stap overgeslagen")
        return 0, 0

    wb = openpyxl.load_workbook(LEDENLIJST, read_only=True, data_only=True)
    ws = wb["Complete Ledenlijst"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    leden = []
    for row in rows[1:]:
        rel = str(row[1]).strip() if row[1] else ""
        naam = str(row[2]).strip() if row[2] else ""
        geslacht = str(row[3]).strip() if row[3] else ""
        gebdat_raw = str(row[4]).strip() if row[4] else ""
        gebdat = None
        if gebdat_raw and gebdat_raw != "None":
            parts = gebdat_raw.split("-")
            if len(parts) == 3 and len(parts[2]) == 4:
                gebdat = f"{parts[2]}-{parts[1]}-{parts[0]}"
        if rel and naam:
            leden.append({"rel": rel, "naam": naam, "geslacht": geslacht,
                          "geboortedatum": gebdat})

    print(f"  Ledenlijst geladen: {len(leden)} leden")

    # Build lookups
    leden_by_gebdat = {}
    for l in leden:
        if l["geboortedatum"]:
            leden_by_gebdat.setdefault(l["geboortedatum"], []).append(l)

    leden_by_key = {}
    for l in leden:
        ach, vn = extract_achternaam_voornaam(l["naam"])
        key = make_match_key(ach, vn)
        if key and key != "|":
            leden_by_key.setdefault(key, []).append(l)
        # Simplified key: first word of voornaam
        if vn:
            first_vn = vn.split()[0] if vn.split() else ""
            if first_vn:
                skey = make_match_key(ach, first_vn)
                if skey and skey != "|":
                    leden_by_key.setdefault(skey, []).append(l)

    # Phase 1: Match OW-spelers naar Sportlink Rel.nr
    ow_ids = [sid for sid, sp in registry.by_rel_nr.items()
              if sp["bron_id"] == "name-match"]
    upgrades = []  # (ow_id, lid)

    for ow_id in ow_ids:
        sp = registry.by_rel_nr[ow_id]
        sp_ach, sp_vn = extract_achternaam_voornaam(sp["naam"])
        roep = sp.get("roepnaam") or ""
        best = None

        # Strategy 1: geboortedatum + achternaam
        if sp["geboortedatum"] and sp["geboortedatum"] in leden_by_gebdat:
            norm_sp_ach = normalize_name(sp_ach)
            for l in leden_by_gebdat[sp["geboortedatum"]]:
                l_ach, _ = extract_achternaam_voornaam(l["naam"])
                norm_l_ach = normalize_name(l_ach)
                # Achternaam moet voldoende overeenkomen
                if (norm_sp_ach == norm_l_ach or
                        norm_sp_ach.split()[0] == norm_l_ach.split()[0]):
                    best = l
                    break

        # Strategy 2: naam-key match (alleen als uniek in ledenlijst)
        if not best:
            for try_vn in [sp_vn, roep]:
                if not try_vn:
                    continue
                key = make_match_key(sp_ach, try_vn)
                if key in leden_by_key:
                    candidates = leden_by_key[key]
                    if len(candidates) == 1:
                        best = candidates[0]
                        break
                    # Meerdere — check geboortedatum als tiebreaker
                    if sp["geboortedatum"]:
                        for c in candidates:
                            if c["geboortedatum"] == sp["geboortedatum"]:
                                best = c
                                break
                    if best:
                        break

        if best and best["rel"] not in registry.by_rel_nr:
            upgrades.append((ow_id, best))

    # Execute upgrades: verplaats OW-speler naar Sportlink Rel.nr
    upgraded = 0
    for ow_id, lid in upgrades:
        sp = registry.by_rel_nr.pop(ow_id)
        sp["speler_id"] = lid["rel"]
        sp["bron_id"] = "sportlink"
        sp["naam"] = lid["naam"]
        if lid["geslacht"]:
            sp["geslacht"] = lid["geslacht"]
        if lid["geboortedatum"]:
            sp["geboortedatum"] = lid["geboortedatum"]
        registry.by_rel_nr[lid["rel"]] = sp
        registry._register_keys(lid["rel"], lid["naam"], sp.get("roepnaam"))
        upgraded += 1

    # Phase 2: merge OW-spelers die nu matchen met bestaande Sportlink-entries
    ow_ids_remaining = [sid for sid, sp in registry.by_rel_nr.items()
                        if sp["bron_id"] == "name-match"]
    merge_count = 0
    for ow_id in ow_ids_remaining:
        sp = registry.by_rel_nr[ow_id]
        sp_ach, sp_vn = extract_achternaam_voornaam(sp["naam"])
        roep = sp.get("roepnaam") or ""
        best = None

        if sp["geboortedatum"] and sp["geboortedatum"] in leden_by_gebdat:
            norm_sp_ach = normalize_name(sp_ach)
            for l in leden_by_gebdat[sp["geboortedatum"]]:
                l_ach, _ = extract_achternaam_voornaam(l["naam"])
                if normalize_name(l_ach).split()[0] == norm_sp_ach.split()[0]:
                    if l["rel"] in registry.by_rel_nr:
                        best = l
                        break

        if best:
            target = registry.by_rel_nr[best["rel"]]
            source = registry.by_rel_nr.pop(ow_id)
            for sz, info in source["seizoenen"].items():
                if sz not in target["seizoenen"]:
                    target["seizoenen"][sz] = info
            merge_count += 1

    # Phase 3: verrijk bestaande spelers met ontbrekende velden
    # 3a: Sportlink-spelers via Relatiecode
    leden_by_rel = {l["rel"]: l for l in leden}
    enriched = 0
    for sp in registry.by_rel_nr.values():
        if sp["speler_id"] in leden_by_rel:
            l = leden_by_rel[sp["speler_id"]]
            changed = False
            if l["geslacht"] and sp["geslacht"] != l["geslacht"]:
                sp["geslacht"] = l["geslacht"]
                changed = True
            if not sp["geboortedatum"] and l["geboortedatum"]:
                sp["geboortedatum"] = l["geboortedatum"]
                changed = True
            if changed:
                enriched += 1

    # 3b: OW-spelers geslacht via naam-matching met ledenlijst
    gender_fixed = 0
    for sp in registry.by_rel_nr.values():
        if not sp["speler_id"].startswith("OW-"):
            continue
        if sp["geslacht"]:
            continue  # al gevuld, niet overschrijven
        sp_ach, sp_vn = extract_achternaam_voornaam(sp["naam"])
        roep = sp.get("roepnaam") or ""
        for try_vn in [sp_vn, roep]:
            if not try_vn:
                continue
            key = make_match_key(sp_ach, try_vn)
            if key in leden_by_key:
                candidates = leden_by_key[key]
                genders = set(c["geslacht"] for c in candidates if c["geslacht"])
                if len(genders) == 1:
                    sp["geslacht"] = genders.pop()
                    gender_fixed += 1
                    break

    print(f"  OW >> Sportlink upgrade:   {upgraded}")
    print(f"  OW >> bestaande gemerged:  {merge_count}")
    print(f"  Velden verrijkt:           {enriched}")
    print(f"  Geslacht via naam-match:   {gender_fixed}")
    return upgraded, merge_count


# --- Diagnostiek: gemiste matches ---------------------------

def diagnose_potential_matches(registry):
    """Zoek potentiële gemiste OW~Sportlink matches via fuzzy matching."""
    print("\n" + "=" * 60)
    print("DIAGNOSTIEK: Potentiële gemiste matches")
    print("=" * 60)

    ow_spelers = {sid: sp for sid, sp in registry.by_rel_nr.items()
                  if sp["bron_id"] == "name-match"}
    sl_spelers = {sid: sp for sid, sp in registry.by_rel_nr.items()
                  if sp["bron_id"] == "sportlink"}

    if not ow_spelers:
        print("  Geen OW-spelers meer — alles gematcht!")
        return []

    # Bouw fuzzy-index voor Sportlink-spelers
    sl_by_fuzzy_ach = defaultdict(list)
    sl_by_voornaam = defaultdict(list)
    for sid, sp in sl_spelers.items():
        ach, vn = extract_achternaam_voornaam(sp["naam"])
        roep = sp.get("roepnaam") or ""
        fuzzy = _fuzzy_achternaam(ach)
        if fuzzy:
            sl_by_fuzzy_ach[fuzzy].append((sid, sp, ach, vn))
        # Index op voornaam/roepnaam
        for v in [vn, roep]:
            if v:
                vn_norm = normalize_name(v.split()[0]) if v.split() else ""
                if vn_norm and len(vn_norm) > 1:
                    sl_by_voornaam[vn_norm].append((sid, sp, ach))

    candidates = []
    for ow_id, ow_sp in ow_spelers.items():
        ow_ach, ow_vn = extract_achternaam_voornaam(ow_sp["naam"])
        ow_roep = ow_sp.get("roepnaam") or ""
        ow_fuzzy = _fuzzy_achternaam(ow_ach)
        ow_seasons = set(ow_sp["seizoenen"].keys())

        matches = []

        # Strategie 1: Fuzzy achternaam + voornaam/roepnaam match
        if ow_fuzzy and ow_fuzzy in sl_by_fuzzy_ach:
            for sl_id, sl_sp, sl_ach, sl_vn in sl_by_fuzzy_ach[ow_fuzzy]:
                sl_roep = sl_sp.get("roepnaam") or ""
                # Check voornaam overlap
                ow_vnames = {normalize_name(v.split()[0]) for v in [ow_vn, ow_roep] if v and v.split()}
                sl_vnames = {normalize_name(v.split()[0]) for v in [sl_vn, sl_roep] if v and v.split()}
                if ow_vnames & sl_vnames:
                    sl_seasons = set(sl_sp["seizoenen"].keys())
                    overlap = ow_seasons & sl_seasons
                    matches.append({
                        "sl_id": sl_id, "sl_naam": sl_sp["naam"],
                        "sl_roep": sl_roep, "sl_gebdat": sl_sp["geboortedatum"],
                        "seizoen_overlap": sorted(overlap),
                        "reden": f"fuzzy achternaam ({ow_ach} ~ {sl_ach})"
                    })

        # Strategie 2: Geboortedatum match + vergelijkbare achternaam (eerste 3 letters)
        if ow_sp["geboortedatum"]:
            for sl_id, sl_sp in sl_spelers.items():
                if sl_sp["geboortedatum"] == ow_sp["geboortedatum"]:
                    sl_ach, sl_vn = extract_achternaam_voornaam(sl_sp["naam"])
                    ow_ach_root = _fuzzy_achternaam(ow_ach)[:3]
                    sl_ach_root = _fuzzy_achternaam(sl_ach)[:3]
                    if ow_ach_root and ow_ach_root == sl_ach_root:
                        # Voorkom duplicaat met strategie 1
                        if not any(m["sl_id"] == sl_id for m in matches):
                            matches.append({
                                "sl_id": sl_id, "sl_naam": sl_sp["naam"],
                                "sl_roep": sl_sp.get("roepnaam", ""),
                                "sl_gebdat": sl_sp["geboortedatum"],
                                "seizoen_overlap": [],
                                "reden": f"geboortedatum + achternaam-start ({ow_ach[:3]}~{sl_ach[:3]})"
                            })

        # Strategie 3: Aangrenzende seizoenen + exacte voornaam
        if not matches:
            for v in [ow_vn, ow_roep]:
                if not v:
                    continue
                vn_norm = normalize_name(v.split()[0]) if v.split() else ""
                if vn_norm and vn_norm in sl_by_voornaam:
                    for sl_id, sl_sp, sl_ach in sl_by_voornaam[vn_norm]:
                        sl_seasons = set(sl_sp["seizoenen"].keys())
                        # Check of seizoenen aangrenzend zijn (bijv. OW stopt 2018, SL start 2019)
                        adjacent = False
                        for ow_s in ow_seasons:
                            ow_end_year = int(ow_s.split("-")[1])
                            for sl_s in sl_seasons:
                                sl_start_year = int(sl_s.split("-")[0])
                                if abs(ow_end_year - sl_start_year) <= 1:
                                    adjacent = True
                                    break
                            if adjacent:
                                break
                        if adjacent:
                            # Voorkom duplicaat
                            if not any(m["sl_id"] == sl_id for m in matches):
                                matches.append({
                                    "sl_id": sl_id, "sl_naam": sl_sp["naam"],
                                    "sl_roep": sl_sp.get("roepnaam", ""),
                                    "sl_gebdat": sl_sp["geboortedatum"],
                                    "seizoen_overlap": [],
                                    "reden": f"aangrenzende seizoenen + voornaam ({v})"
                                })

        if matches:
            candidates.append((ow_id, ow_sp, matches))

    # Print resultaten
    if not candidates:
        print("  Geen potentiële matches gevonden.")
    else:
        print(f"\n  {len(candidates)} OW-spelers met potentiële Sportlink-match:\n")
        for ow_id, ow_sp, matches in candidates:
            ow_seasons = sorted(ow_sp["seizoenen"].keys())
            print(f"  {ow_id}: {ow_sp['naam']} (roep: {ow_sp.get('roepnaam','')}, "
                  f"geb: {ow_sp.get('geboortedatum','?')}, "
                  f"seizoenen: {', '.join(ow_seasons)})")
            for m in matches:
                overlap_str = f", overlap: {m['seizoen_overlap']}" if m['seizoen_overlap'] else ""
                print(f"    >>{m['sl_id']}: {m['sl_naam']} (roep: {m['sl_roep']}, "
                      f"geb: {m['sl_gebdat']}{overlap_str})")
                print(f"      reden: {m['reden']}")
            print()

    return candidates


# --- Hertelling ---------------------------------------------

def validate_season_counts(spelers_list):
    """Hertelling: M/V per team per seizoen, opslaan als hertelling.json."""
    print("\n" + "=" * 60)
    print("HERTELLING PER SEIZOEN")
    print("=" * 60)

    hertelling = {}
    all_seizoenen = sorted(set(
        sz for sp in spelers_list for sz in sp["seizoenen"]
    ))

    for seizoen in all_seizoenen:
        teams = defaultdict(lambda: {"M": 0, "V": 0, "onbekend": 0})
        totaal_m = 0
        totaal_v = 0
        totaal = 0

        for sp in spelers_list:
            if seizoen not in sp["seizoenen"]:
                continue
            sz_info = sp["seizoenen"][seizoen]
            team = sz_info.get("team", "?") or "?"
            geslacht = sp.get("geslacht", "")
            totaal += 1
            if geslacht == "M":
                teams[team]["M"] += 1
                totaal_m += 1
            elif geslacht == "V":
                teams[team]["V"] += 1
                totaal_v += 1
            else:
                teams[team]["onbekend"] += 1

        # Sorteer teams
        team_dict = {}
        for t in sorted(teams.keys()):
            entry = {"M": teams[t]["M"], "V": teams[t]["V"]}
            if teams[t]["onbekend"] > 0:
                entry["onbekend"] = teams[t]["onbekend"]
            team_dict[t] = entry

        hertelling[seizoen] = {
            "totaal": totaal,
            "M": totaal_m,
            "V": totaal_v,
            "teams": team_dict,
            "aantal_teams": len(team_dict),
        }

        print(f"\n  {seizoen}: {totaal} spelers ({totaal_m}M / {totaal_v}V), {len(team_dict)} teams")
        for t in sorted(teams.keys()):
            t_data = teams[t]
            parts = []
            if t_data["M"]: parts.append(f"{t_data['M']}M")
            if t_data["V"]: parts.append(f"{t_data['V']}V")
            if t_data["onbekend"]: parts.append(f"{t_data['onbekend']}?")
            print(f"    {t:10s} {' + '.join(parts):>12s} = {sum(t_data.values()):>3d}")

    # Opslaan
    out_path = os.path.join(OUT, "hertelling.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(hertelling, f, ensure_ascii=False, indent=2)
    print(f"\n  >> {out_path}")

    return hertelling


# --- Main ---------------------------------------------------

def main():
    registry = SpelerRegistry()

    print("=" * 60)
    print("STAP 1: Supersheets 2021-2026 (met Rel.nr)")
    print("=" * 60)
    parse_supersheets(registry)
    print(f"\n  >> {len(registry.by_rel_nr)} spelers na Supersheets")

    print("\n" + "=" * 60)
    print("STAP 2: A2 formulieren 2018-2022 (met Rel.nr)")
    print("=" * 60)
    parse_all_a2(registry)
    print(f"\n  >> {len(registry.by_rel_nr)} spelers na A2")

    print("\n" + "=" * 60)
    print("STAP 3: Opstellingen met geboortedatum (2017-2020)")
    print("=" * 60)
    xlsx_files = [
        ("2017-2018", os.path.join(BASE, "Opstelling 20172018.xlsx")),
        ("2019-2020", os.path.join(BASE, "Opstelling 20192020.xlsx")),
    ]
    for seizoen, path in xlsx_files:
        parse_xlsx_opstelling(registry, seizoen, path)
    print(f"\n  >> {len(registry.by_rel_nr)} spelers na 2017-2020")

    print("\n" + "=" * 60)
    print("STAP 3b: Teamindeling 2020-2021 (COVID-seizoen)")
    print("=" * 60)
    parse_xlsx_opstelling(
        registry, "2020-2021",
        os.path.join(BASE, "Teamindeling seizoen 20-21.xlsx"),
        include_sheets=["Sen", "A", "B", "C", "D", "E", "F"],
    )
    print(f"\n  >> {len(registry.by_rel_nr)} spelers na 2020-2021")

    print("\n" + "=" * 60)
    print("STAP 4: Opstellingen zonder geboortedatum (2010-2016)")
    print("=" * 60)
    xlsx_old = [
        ("2012-2013", os.path.join(BASE, "Opstelling 20122013.xlsx")),
        ("2016-2017", os.path.join(BASE, "Opstelling 20162017.xlsx")),
    ]
    for seizoen, path in xlsx_old:
        parse_xlsx_opstelling(registry, seizoen, path)
    xls_files = [
        ("2010-2011", os.path.join(BASE, "Opstelling 20102011.xls")),
        ("2011-2012", os.path.join(BASE, "Opstelling 20112012.xls")),
        ("2013-2014", os.path.join(BASE, "opstelling 20132014.xls")),
        ("2014-2015", os.path.join(BASE, "opstelling 20142015.xls")),
        ("2015-2016", os.path.join(BASE, "opstelling 20152016.xls")),
    ]
    for seizoen, path in xls_files:
        parse_xls_opstelling(registry, seizoen, path)
    print(f"\n  >> {len(registry.by_rel_nr)} spelers totaal")

    print("\n" + "=" * 60)
    print("STAP 5: Deduplicatie")
    print("=" * 60)
    merged = merge_duplicates(registry)
    print(f"  {merged} duplicaten samengevoegd")
    print(f"  >> {len(registry.by_rel_nr)} spelers na deduplicatie")

    print("\n" + "=" * 60)
    print("STAP 6: Verrijking met Complete Ledenlijst")
    print("=" * 60)
    upgraded, ow_merged = enrich_from_ledenlijst(registry)
    print(f"  >> {len(registry.by_rel_nr)} spelers na verrijking")

    # Tweede merge-pass: sommige MERGE_TABLE entries verwijzen naar Sportlink IDs
    # die pas na ledenlijst-verrijking bestaan (bijv. Ooijen/Ooyen)
    print("\n" + "=" * 60)
    print("STAP 6b: Deduplicatie (tweede pass na verrijking)")
    print("=" * 60)
    merged2 = merge_duplicates(registry)
    merged += merged2
    print(f"  {merged2} extra duplicaten samengevoegd")
    print(f"  >> {len(registry.by_rel_nr)} spelers na tweede deduplicatie")

    # -- Diagnostiek: gemiste matches --
    diagnose_potential_matches(registry)

    # -- Statistieken --
    print("\n" + "=" * 60)
    print("STATISTIEKEN")
    print("=" * 60)
    sportlink = sum(1 for s in registry.by_rel_nr.values() if s["bron_id"] == "sportlink")
    matched = sum(1 for s in registry.by_rel_nr.values() if s["bron_id"] == "name-match")
    with_gebdat = sum(1 for s in registry.by_rel_nr.values() if s["geboortedatum"])
    gestopt = sum(1 for s in registry.by_rel_nr.values() if s["gestopt"])

    print(f"  Totaal spelers: {len(registry.by_rel_nr)}")
    print(f"  Met Sportlink Rel.nr: {sportlink}")
    print(f"  Met OW-code (name-match): {matched}")
    print(f"  Met geboortedatum: {with_gebdat}")
    print(f"  Gestopt: {gestopt}")

    season_counts = defaultdict(int)
    for sp in registry.by_rel_nr.values():
        season_counts[len(sp["seizoenen"])] += 1
    print(f"\n  Aantal seizoenen per speler:")
    for n in sorted(season_counts.keys()):
        print(f"    {n} seizoen(en): {season_counts[n]} spelers")

    all_seizoenen = sorted(set(
        s for sp in registry.by_rel_nr.values() for s in sp["seizoenen"]
    ))
    print(f"\n  Spelers per seizoen:")
    for sz in all_seizoenen:
        n = sum(1 for sp in registry.by_rel_nr.values() if sz in sp["seizoenen"])
        print(f"    {sz}: {n} spelers")

    # -- Opslaan --
    print("\n" + "=" * 60)
    print("OPSLAAN")
    print("=" * 60)
    os.makedirs(OUT, exist_ok=True)

    # Filter: verwijder spelers zonder seizoenen (bestuur/vrijwilligers zonder team)
    all_spelers = list(registry.by_rel_nr.values())
    no_seasons = [s for s in all_spelers if len(s["seizoenen"]) == 0]
    if no_seasons:
        print(f"\n  Gefilterd: {len(no_seasons)} leden zonder teamtoewijzing verwijderd")
    spelers_list = sorted(
        [s for s in all_spelers if len(s["seizoenen"]) > 0],
        key=lambda s: s["speler_id"]
    )
    for sp in spelers_list:
        sp["seizoenen"] = dict(sorted(sp["seizoenen"].items()))

    out_path = os.path.join(OUT, "spelerspaden.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(spelers_list, f, ensure_ascii=False, indent=2)
    print(f"  >> {out_path} ({len(spelers_list)} spelers)")

    # -- Hertelling --
    validate_season_counts(spelers_list)

    # Statistieken op de DEFINITIEVE (gefilterde) lijst
    final_sportlink = sum(1 for s in spelers_list if s["bron_id"] == "sportlink")
    final_ow = sum(1 for s in spelers_list if s["bron_id"] == "name-match")
    final_gebdat = sum(1 for s in spelers_list if s["geboortedatum"])
    final_gestopt = sum(1 for s in spelers_list if s["gestopt"])
    final_seizoenen = sorted(set(
        s for sp in spelers_list for s in sp["seizoenen"]
    ))

    meta = {
        "gegenereerd": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "bronbestanden": [
            "Supersheet indeling 1.0 - 2021-2022.xlsx",
            "Supersheet indeling 1.0 - 2022-2023.xlsx",
            "Supersheet indeling 1.0 - 2023-2024.xlsx",
            "Supersheet indeling 1.0 - 2024-2025.xlsx",
            "Supersheet indeling 1.0 - 2025-2026.xlsx",
            "A2 zwt - Oranje Wit (D) 2018 2019.xlsm",
            "A2 zwt - Oranje Wit (D) 06-2020.xlsm",
            "A2 mix - Oranje Wit (D)  09-2021.xlsm",
            "A2 mix - Oranje Wit (D) 20-2022.xlsm",
            "Opstelling 20172018.xlsx",
            "Opstelling 20192020.xlsx",
            "Teamindeling seizoen 20-21.xlsx",
            "Opstelling 20122013.xlsx",
            "Opstelling 20162017.xlsx",
            "Opstelling 20102011.xls",
            "Opstelling 20112012.xls",
            "opstelling 20132014.xls",
            "opstelling 20142015.xls",
            "opstelling 20152016.xls",
            "Complete_Ledenlijst_TC.xlsx",
        ],
        "seizoenen": final_seizoenen,
        "statistieken": {
            "totaal_spelers": len(spelers_list),
            "met_sportlink_id": final_sportlink,
            "met_ow_code": final_ow,
            "met_geboortedatum": final_gebdat,
            "gestopt": final_gestopt,
            "duplicaten_gemerged": merged,
            "leden_zonder_team_gefilterd": len(no_seasons),
        }
    }
    meta_path = os.path.join(OUT, "spelerspaden-meta.json")
    with open(meta_path, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    print(f"  >> {meta_path}")


if __name__ == "__main__":
    main()
